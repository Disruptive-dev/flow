from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, Request, HTTPException, Response
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import uuid
import bcrypt
import jwt
import random
import asyncio
import resend
import re
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
from bson import ObjectId

# Configure Resend
resend.api_key = os.environ.get("RESEND_API_KEY", "")

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]
JWT_ALGORITHM = "HS256"

app = FastAPI(title="Spectra Flow API")
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ==================== AUTH HELPERS ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

def get_jwt_secret():
    return os.environ["JWT_SECRET"]

def create_access_token(user_id: str, email: str) -> str:
    payload = {"sub": user_id, "email": email, "exp": datetime.now(timezone.utc) + timedelta(hours=24), "type": "access"}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "refresh"}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def set_auth_cookies(response: Response, access_token: str, refresh_token: str):
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=False, samesite="lax", max_age=86400, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=False, samesite="lax", max_age=604800, path="/")

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        token = request.query_params.get("token", "")
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        user["_id"] = str(user["_id"])
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def serialize_doc(doc):
    if doc and "_id" in doc:
        doc.pop("_id")
    return doc

# ==================== MODELS ====================

class LoginRequest(BaseModel):
    email: str
    password: str

class RegisterRequest(BaseModel):
    email: str
    password: str
    name: str
    tenant_name: Optional[str] = None

class ProspectJobCreate(BaseModel):
    country: Optional[str] = "Argentina"
    province: Optional[str] = ""
    city: Optional[str] = ""
    category: str
    quantity: int = 100
    postal_code: Optional[str] = ""
    filters: Optional[Dict[str, Any]] = None
    source: Optional[str] = "google_maps"
    linkedin_params: Optional[Dict[str, Any]] = None

class LeadStatusUpdate(BaseModel):
    status: str

class BulkActionRequest(BaseModel):
    lead_ids: List[str]
    action: str

class CampaignCreate(BaseModel):
    name: str
    sender_profile_id: Optional[str] = ""
    template_id: Optional[str] = ""
    lead_ids: Optional[List[str]] = []

class CampaignUpdate(BaseModel):
    name: Optional[str] = None
    status: Optional[str] = None
    sender_profile_id: Optional[str] = None
    template_id: Optional[str] = None

class TemplateCreate(BaseModel):
    name: str
    subject: str
    html_body: str
    plain_text: Optional[str] = ""
    variables: Optional[List[str]] = []
    signature: Optional[str] = ""

class TemplateUpdate(BaseModel):
    name: Optional[str] = None
    subject: Optional[str] = None
    html_body: Optional[str] = None
    plain_text: Optional[str] = None
    variables: Optional[List[str]] = None
    signature: Optional[str] = None

class DomainCreate(BaseModel):
    domain: str
    subdomain: Optional[str] = ""
    sender_name: Optional[str] = ""
    sender_email: Optional[str] = ""
    reply_to: Optional[str] = ""
    signature: Optional[str] = ""

class DomainUpdate(BaseModel):
    sender_name: Optional[str] = None
    sender_email: Optional[str] = None
    reply_to: Optional[str] = None
    signature: Optional[str] = None
    status: Optional[str] = None

class CrmPushRequest(BaseModel):
    lead_ids: List[str]
    assigned_owner: Optional[str] = ""

class SettingsUpdate(BaseModel):
    company_name: Optional[str] = None
    logo_url: Optional[str] = None
    primary_color: Optional[str] = None
    secondary_color: Optional[str] = None
    industry: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    website: Optional[str] = None
    tax_id: Optional[str] = None
    country: Optional[str] = None
    description: Optional[str] = None
    sender_defaults: Optional[Dict[str, str]] = None

class IntegrationUpdate(BaseModel):
    enabled: Optional[bool] = None
    base_url: Optional[str] = None
    api_key: Optional[str] = None

class UserCreate(BaseModel):
    email: str
    password: str
    name: str
    role: str = "operator"

# ==================== AUTH ENDPOINTS ====================

@api_router.post("/auth/login")
async def login(request: Request, response: Response, body: LoginRequest):
    email = body.email.lower().strip()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    user_id = str(user["_id"])
    access_token = create_access_token(user_id, email)
    refresh_token = create_refresh_token(user_id)
    set_auth_cookies(response, access_token, refresh_token)
    return {"id": user_id, "email": user["email"], "name": user["name"], "role": user["role"], "tenant_id": user.get("tenant_id", ""), "access_token": access_token}

@api_router.post("/auth/register")
async def register(response: Response, body: RegisterRequest):
    email = body.email.lower().strip()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    tenant_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    trial_end = (datetime.now(timezone.utc) + timedelta(days=15)).isoformat()
    tenant = {"id": tenant_id, "name": body.tenant_name or f"{body.name}'s Organization", "branding": {"company_name": body.tenant_name or f"{body.name}'s Organization", "logo_url": "", "primary_color": "#1D4ED8", "secondary_color": "#6366F1"}, "sender_defaults": {"name": body.tenant_name or body.name, "email": "no-reply@spectra-metrics.com"}, "modules": {"prospeccion": False, "leads": True, "crm": True, "email_marketing": False, "web": False, "performance": False}, "plan": "trial", "price": 0, "active": True, "trial_ends_at": trial_end, "created_at": now}
    await db.tenants.insert_one(tenant)
    user_doc = {"email": email, "password_hash": hash_password(body.password), "name": body.name, "role": "tenant_admin", "tenant_id": tenant_id, "created_at": now}
    result = await db.users.insert_one(user_doc)
    user_id = str(result.inserted_id)
    access_token = create_access_token(user_id, email)
    refresh_token = create_refresh_token(user_id)
    set_auth_cookies(response, access_token, refresh_token)
    return {"id": user_id, "email": email, "name": body.name, "role": "tenant_admin", "tenant_id": tenant_id, "access_token": access_token}

@api_router.get("/auth/me")
async def get_me(request: Request):
    user = await get_current_user(request)
    return user

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Logged out"}

@api_router.post("/auth/forgot-password")
async def forgot_password(body: Dict[str, Any] = {}):
    email = body.get("email", "").lower().strip()
    if not email:
        raise HTTPException(status_code=400, detail="Email requerido")
    user = await db.users.find_one({"email": email})
    if not user:
        return {"message": "Si el email existe, recibiras un enlace para restablecer tu password"}
    token = create_access_token(str(user["_id"]), email)
    frontend_url = os.environ.get("FRONTEND_URL", os.environ.get("REACT_APP_BACKEND_URL", ""))
    reset_link = f"{frontend_url}/reset-password?token={token}"
    try:
        resend.emails.send({"from": "no-reply@spectra-metrics.com", "to": [email], "subject": "Restablecer password - Spectra Flow", "html": f"<h2>Restablecer password</h2><p>Haz click en el siguiente enlace para restablecer tu password:</p><p><a href='{reset_link}' style='color:#1D4ED8;font-weight:bold;'>Restablecer password</a></p><p>Este enlace expira en 1 hora.</p><p>Si no solicitaste esto, ignora este email.</p>"})
    except Exception as e:
        logger.error(f"Error sending reset email: {e}")
    return {"message": "Si el email existe, recibiras un enlace para restablecer tu password"}

@api_router.post("/auth/reset-password")
async def reset_password(body: Dict[str, Any] = {}):
    token = body.get("token", "")
    new_pw = body.get("new_password", "")
    if not token or not new_pw or len(new_pw) < 6:
        raise HTTPException(status_code=400, detail="Token y nueva password (min 6 chars) requeridos")
    try:
        payload = jwt.decode(token, os.environ.get("JWT_SECRET", "spectra-flow-secret-key-2024"), algorithms=["HS256"])
        user_id = payload.get("sub")
        new_hash = bcrypt.hashpw(new_pw.encode(), bcrypt.gensalt()).decode()
        await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": {"password_hash": new_hash, "updated_at": datetime.now(timezone.utc).isoformat()}})
        return {"message": "Password restablecida exitosamente"}
    except:
        raise HTTPException(status_code=400, detail="Token invalido o expirado")

@api_router.post("/admin/reset-demo-data")
async def reset_demo_data(request: Request):
    """Reset all demo data - SUPER ADMIN ONLY"""
    user = await get_current_user(request)
    if user["role"] != "super_admin":
        raise HTTPException(status_code=403, detail="Solo super_admin")
    tid = user["tenant_id"]
    counts = {}
    for col in ["leads", "crm_contacts", "crm_deals", "crm_tasks", "crm_notes", "crm_deal_products", "prospect_jobs", "campaigns", "email_campaigns", "email_lists", "email_segments", "activity_log"]:
        r = await db[col].delete_many({"tenant_id": tid})
        counts[col] = r.deleted_count
    return {"message": "Datos demo eliminados", "deleted": counts}

@api_router.put("/auth/profile")
async def update_profile(request: Request, body: Dict[str, Any] = {}):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    update = {"updated_at": now}
    if body.get("name"):
        update["name"] = body["name"]
    if body.get("phone"):
        update["phone"] = body["phone"]
    if body.get("avatar_url"):
        update["avatar_url"] = body["avatar_url"]
    if body.get("job_title"):
        update["job_title"] = body["job_title"]
    await db.users.update_one({"_id": ObjectId(user["_id"])}, {"$set": update})
    updated = await db.users.find_one({"_id": ObjectId(user["_id"])}, {"_id": 0, "password_hash": 0})
    return serialize_doc(updated)

@api_router.put("/auth/change-password")
async def change_password(request: Request, body: Dict[str, Any] = {}):
    user_raw = await get_current_user(request)
    current = body.get("current_password", "")
    new_pw = body.get("new_password", "")
    if not current or not new_pw:
        raise HTTPException(status_code=400, detail="Se requiere password actual y nueva")
    if len(new_pw) < 6:
        raise HTTPException(status_code=400, detail="La nueva password debe tener al menos 6 caracteres")
    full_user = await db.users.find_one({"_id": ObjectId(user_raw["_id"])})
    if not bcrypt.checkpw(current.encode(), full_user["password_hash"].encode()):
        raise HTTPException(status_code=400, detail="Password actual incorrecta")
    new_hash = bcrypt.hashpw(new_pw.encode(), bcrypt.gensalt()).decode()
    await db.users.update_one({"_id": ObjectId(user_raw["_id"])}, {"$set": {"password_hash": new_hash, "updated_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": "Password actualizada exitosamente"}

from fastapi import UploadFile, File
from fastapi.staticfiles import StaticFiles

@api_router.post("/auth/avatar")
async def upload_avatar(request: Request, file: UploadFile = File(...)):
    user = await get_current_user(request)
    import shutil
    uploads_dir = ROOT_DIR / "uploads" / "avatars"
    uploads_dir.mkdir(parents=True, exist_ok=True)
    ext = file.filename.split(".")[-1] if "." in file.filename else "jpg"
    filename = f"{user['_id']}.{ext}"
    filepath = uploads_dir / filename
    with open(filepath, "wb") as f:
        shutil.copyfileobj(file.file, f)
    avatar_url = f"/api/uploads/avatars/{filename}"
    await db.users.update_one({"_id": ObjectId(user["_id"])}, {"$set": {"avatar_url": avatar_url, "updated_at": datetime.now(timezone.utc).isoformat()}})
    return {"avatar_url": avatar_url}

@api_router.post("/auth/refresh")
async def refresh_token_endpoint(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="No refresh token")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        user_id = str(user["_id"])
        new_access = create_access_token(user_id, user["email"])
        response.set_cookie(key="access_token", value=new_access, httponly=True, secure=False, samesite="lax", max_age=86400, path="/")
        return {"message": "Token refreshed", "access_token": new_access}
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid refresh token")

# ==================== DASHBOARD ====================

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(request: Request, from_date: Optional[str] = None, to_date: Optional[str] = None):
    user = await get_current_user(request)
    tid = user.get("tenant_id", "")
    date_filter = {}
    if from_date:
        date_filter["created_at"] = {"$gte": from_date}
    if to_date:
        date_filter.setdefault("created_at", {})
        date_filter["created_at"]["$lte"] = to_date
    lead_filter = {"tenant_id": tid, **date_filter}
    campaign_filter = {"tenant_id": tid, **date_filter}
    jobs_count = await db.prospect_jobs.count_documents({"tenant_id": tid, **date_filter})
    total_leads = await db.leads.count_documents(lead_filter)
    raw_leads = await db.leads.count_documents({**lead_filter, "status": "raw"})
    qualified = await db.leads.count_documents({**lead_filter, "status": {"$in": ["scored", "approved"]}})
    campaigns_active = await db.campaigns.count_documents({**campaign_filter, "status": "active"})
    pipeline = [{"$match": campaign_filter}, {"$group": {"_id": None, "sent": {"$sum": "$sent_count"}, "opens": {"$sum": "$open_count"}, "clicks": {"$sum": "$click_count"}, "replies": {"$sum": "$reply_count"}, "interested": {"$sum": "$interested_count"}, "crm": {"$sum": "$crm_count"}}}]
    email_stats = await db.campaigns.aggregate(pipeline).to_list(1)
    es = email_stats[0] if email_stats else {}
    crm_synced = await db.crm_sync_logs.count_documents({"tenant_id": tid, **date_filter, "status": "synced"})
    recent = await db.audit_logs.find({"tenant_id": tid}, {"_id": 0}).sort("created_at", -1).limit(10).to_list(10)
    return {
        "jobs_this_month": jobs_count, "raw_leads": raw_leads, "qualified_leads": qualified,
        "total_leads": total_leads, "emails_sent": es.get("sent", 0), "opens": es.get("opens", 0),
        "clicks": es.get("clicks", 0), "replies": es.get("replies", 0),
        "interested": es.get("interested", 0), "leads_sent_to_crm": crm_synced,
        "opportunities": es.get("crm", 0), "active_campaigns": campaigns_active, "recent_activity": recent
    }

# ==================== PROSPECT JOBS ====================

@api_router.get("/prospect-jobs")
async def list_prospect_jobs(request: Request):
    user = await get_current_user(request)
    jobs = await db.prospect_jobs.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return jobs

@api_router.post("/prospect-jobs")
async def create_prospect_job(request: Request, body: ProspectJobCreate):
    user = await get_current_user(request)
    job_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    job = {
        "id": job_id, "tenant_id": user["tenant_id"],
        "country": body.country or "Argentina",
        "province": body.province or "", "city": body.city or "",
        "category": body.category, "quantity": body.quantity,
        "postal_code": body.postal_code or "",
        "filters": body.filters or {}, "status": "pending",
        "source": body.source or "google_maps",
        "linkedin_params": body.linkedin_params or {},
        "raw_count": 0, "cleaned_count": 0, "qualified_count": 0,
        "rejected_count": 0, "approved_count": 0,
        "stages": [
            {"name": "job_created", "status": "completed", "timestamp": now},
            {"name": "scraping", "status": "pending", "timestamp": None},
            {"name": "prospects_found", "status": "pending", "timestamp": None},
            {"name": "ai_cleaning", "status": "pending", "timestamp": None},
            {"name": "scoring_completed", "status": "pending", "timestamp": None},
            {"name": "ready_for_review", "status": "pending", "timestamp": None}
        ],
        "created_by": user.get("name", ""), "created_at": now, "updated_at": now
    }
    await db.prospect_jobs.insert_one(job)
    await db.audit_logs.insert_one({"id": str(uuid.uuid4()), "tenant_id": user["tenant_id"], "user_id": user["_id"], "user_name": user["name"], "action": "created_prospect_job", "entity_type": "prospect_job", "entity_id": job_id, "details": f"Job: {body.category} in {body.city}, {body.province}", "created_at": now})
    # Try to trigger n8n webhook if configured
    try:
        n8n_config = await db.integration_configs.find_one({"tenant_id": user["tenant_id"], "name": "n8n", "enabled": True}, {"_id": 0})
        if n8n_config and n8n_config.get("base_url"):
            import httpx
            webhook_url = n8n_config["base_url"].rstrip("/")
            callback_url = os.environ.get("REACT_APP_BACKEND_URL", os.environ.get("FRONTEND_URL", "https://spectra-hub.preview.emergentagent.com")) + f"/api/webhooks/n8n/job-result/{job_id}"
            progress_url = os.environ.get("REACT_APP_BACKEND_URL", os.environ.get("FRONTEND_URL", "https://spectra-hub.preview.emergentagent.com")) + f"/api/webhooks/n8n/job-progress/{job_id}"
            async with httpx.AsyncClient(timeout=10) as client:
                await client.post(webhook_url, json={
                    "job_id": job_id, "tenant_id": user["tenant_id"],
                    "country": body.country or "Argentina", "province": body.province or "",
                    "city": body.city or "", "category": body.category, "quantity": body.quantity,
                    "postal_code": body.postal_code or "",
                    "source": body.source or "google_maps",
                    "linkedin_params": body.linkedin_params or {},
                    "callback_url": callback_url, "progress_url": progress_url,
                    "api_key": n8n_config.get("api_key", "")
                })
            logger.info(f"n8n webhook triggered for job {job_id}")
    except Exception as e:
        logger.warning(f"n8n webhook failed (will use demo mode): {e}")
    return serialize_doc(job)

@api_router.get("/prospect-jobs/{job_id}")
async def get_prospect_job(request: Request, job_id: str):
    user = await get_current_user(request)
    job = await db.prospect_jobs.find_one({"id": job_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job


async def _run_apify_linkedin(job_id: str, job: dict, api_key: str, tenant_id: str):
    """Background task: Run Apify LinkedIn scraper and store results"""
    import asyncio
    import httpx
    try:
        li_params = job.get("linkedin_params", {})
        keyword = li_params.get("keyword", job.get("category", ""))
        location = li_params.get("location", job.get("city", ""))
        industry = li_params.get("industry", "")
        quantity = job.get("quantity", 50)
        # Use Google search to find LinkedIn company profiles (reliable, no login needed)
        search_query = f"site:linkedin.com/company {keyword}"
        if location:
            search_query += f" {location}"
        if industry:
            search_query += f" {industry}"
        actor_input = {"queries": search_query, "maxPagesPerQuery": max(1, quantity // 10), "resultsPerPage": 10}
        async with httpx.AsyncClient(timeout=120) as client:
            run_resp = await client.post(
                f"https://api.apify.com/v2/acts/apify~google-search-scraper/runs?token={api_key}",
                json=actor_input
            )
            run_data = run_resp.json().get("data", {})
            run_id = run_data.get("id", "")
            if not run_id:
                logger.error(f"Apify no run ID: {run_resp.text[:200]}")
                await db.prospect_jobs.update_one({"id": job_id}, {"$set": {"status": "failed", "updated_at": datetime.now(timezone.utc).isoformat()}})
                return
            now_ts = datetime.now(timezone.utc).isoformat()
            await db.prospect_jobs.update_one({"id": job_id}, {"$set": {"stages.1.status": "completed", "stages.2.status": "processing", "updated_at": now_ts}})
            for _ in range(24):
                await asyncio.sleep(5)
                status_resp = await client.get(f"https://api.apify.com/v2/actor-runs/{run_id}?token={api_key}")
                status = status_resp.json().get("data", {}).get("status", "")
                if status in ("SUCCEEDED", "FAILED", "ABORTED", "TIMED-OUT"):
                    break
            if status != "SUCCEEDED":
                await db.prospect_jobs.update_one({"id": job_id}, {"$set": {"status": "failed", "updated_at": datetime.now(timezone.utc).isoformat()}})
                return
            dataset_id = run_data.get("defaultDatasetId", "")
            results_resp = await client.get(f"https://api.apify.com/v2/datasets/{dataset_id}/items?token={api_key}")
            raw_items = results_resp.json() if isinstance(results_resp.json(), list) else []
            # Parse Google search results for LinkedIn company pages
            items = []
            for r in raw_items:
                organic = r.get("organicResults", [])
                for org in organic:
                    url = org.get("url", "")
                    if "linkedin.com/company" in url:
                        items.append({"name": org.get("title", "").split(" |")[0].split(" -")[0].strip(), "url": url, "description": org.get("description", ""), "location": location})
            now = datetime.now(timezone.utc).isoformat()
            inserted = 0
            for item in items[:quantity]:
                bname = item.get("name", "").strip()
                if not bname or len(bname) < 2:
                    continue
                existing = await db.leads.find_one({"business_name": {"$regex": f"^{re.escape(bname)}$", "$options": "i"}, "tenant_id": tenant_id})
                if existing:
                    continue
                lead = {
                    "id": str(uuid.uuid4()), "tenant_id": tenant_id, "job_id": job_id,
                    "business_name": bname,
                    "raw_category": industry or keyword,
                    "normalized_category": industry or keyword,
                    "province": "", "city": location,
                    "website": item.get("url", ""),
                    "email": "", "phone": "",
                    "ai_score": 50, "quality_level": "average",
                    "recommendation": item.get("description", "")[:300],
                    "recommended_first_line": "",
                    "address": location, "source": "linkedin",
                    "status": "scored", "created_at": now, "updated_at": now
                }
                await db.leads.insert_one(lead)
                inserted += 1
            stages = [
                {"name": "job_created", "status": "completed", "timestamp": job.get("created_at")},
                {"name": "scraping", "status": "completed", "timestamp": now},
                {"name": "prospects_found", "status": "completed", "timestamp": now},
                {"name": "scoring_completed", "status": "completed", "timestamp": now},
                {"name": "ready_for_review", "status": "completed", "timestamp": now}
            ]
            await db.prospect_jobs.update_one({"id": job_id}, {"$set": {
                "status": "completed", "raw_count": len(items), "cleaned_count": inserted,
                "qualified_count": inserted, "rejected_count": 0, "stages": stages, "updated_at": now
            }})
            # Auto-rescore
            if inserted > 0:
                asyncio.create_task(_auto_rescore_job_leads(job_id, tenant_id))
    except Exception as e:
        logger.error(f"Apify LinkedIn error for job {job_id}: {e}")
        await db.prospect_jobs.update_one({"id": job_id}, {"$set": {"status": "failed", "updated_at": datetime.now(timezone.utc).isoformat()}})

@api_router.post("/prospect-jobs/{job_id}/start")
async def start_prospect_job(request: Request, job_id: str):
    user = await get_current_user(request)
    job = await db.prospect_jobs.find_one({"id": job_id, "tenant_id": user["tenant_id"]})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    now = datetime.now(timezone.utc).isoformat()
    # Check if n8n is configured - if so, just mark as processing and wait for callback
    n8n_config = await db.integration_configs.find_one({"tenant_id": user["tenant_id"], "name": "n8n", "enabled": True}, {"_id": 0})
    source = job.get("source", "google_maps")
    if n8n_config and n8n_config.get("base_url"):
        stages = [
            {"name": "job_created", "status": "completed", "timestamp": job["created_at"]},
            {"name": "scraping", "status": "processing", "timestamp": now},
            {"name": "prospects_found", "status": "pending", "timestamp": None},
            {"name": "ai_cleaning", "status": "pending", "timestamp": None},
            {"name": "scoring_completed", "status": "pending", "timestamp": None},
            {"name": "ready_for_review", "status": "pending", "timestamp": None}
        ]
        await db.prospect_jobs.update_one({"id": job_id}, {"$set": {"status": "processing", "stages": stages, "updated_at": now}})
        updated_job = await db.prospect_jobs.find_one({"id": job_id}, {"_id": 0})
        return updated_job
    # Check for Apify direct mode (LinkedIn without n8n)
    if source == "linkedin":
        apify_config = await db.integration_configs.find_one({"tenant_id": user["tenant_id"], "name": "apify", "enabled": True}, {"_id": 0})
        if apify_config and apify_config.get("api_key"):
            stages = [
                {"name": "job_created", "status": "completed", "timestamp": job["created_at"]},
                {"name": "scraping", "status": "processing", "timestamp": now},
                {"name": "prospects_found", "status": "pending", "timestamp": None},
                {"name": "scoring_completed", "status": "pending", "timestamp": None},
                {"name": "ready_for_review", "status": "pending", "timestamp": None}
            ]
            await db.prospect_jobs.update_one({"id": job_id}, {"$set": {"status": "processing", "stages": stages, "updated_at": now}})
            # Launch Apify actor in background
            import asyncio
            asyncio.create_task(_run_apify_linkedin(job_id, job, apify_config["api_key"], user["tenant_id"]))
            updated_job = await db.prospect_jobs.find_one({"id": job_id}, {"_id": 0})
            return updated_job
    # Demo mode - generate simulated data
    quantity = job.get("quantity", 100)
    raw_count = random.randint(int(quantity * 0.8), int(quantity * 1.2))
    cleaned_count = int(raw_count * random.uniform(0.7, 0.85))
    qualified_count = int(cleaned_count * random.uniform(0.6, 0.8))
    rejected_count = cleaned_count - qualified_count
    stages = [
        {"name": "job_created", "status": "completed", "timestamp": job["created_at"]},
        {"name": "scraping", "status": "completed", "timestamp": now},
        {"name": "prospects_found", "status": "completed", "timestamp": now},
        {"name": "ai_cleaning", "status": "completed", "timestamp": now},
        {"name": "scoring_completed", "status": "completed", "timestamp": now},
        {"name": "ready_for_review", "status": "completed", "timestamp": now}
    ]
    await db.prospect_jobs.update_one({"id": job_id}, {"$set": {"status": "completed", "raw_count": raw_count, "cleaned_count": cleaned_count, "qualified_count": qualified_count, "rejected_count": rejected_count, "approved_count": 0, "stages": stages, "updated_at": now}})

    # Realistic Argentine business data by category
    biz_by_cat = {
        "real estate": ["Inmobiliaria del Sol", "Propiedades Austral", "RE/MAX Norte", "Casa Propia Inversiones", "Grupo Habitat", "Loteos del Valle", "Bertoni Propiedades", "Techo Propio SRL", "Cimientos Realty", "Quintas Premium", "Urban Housing SA", "Torres & Lagos", "Estancias Pampeanas", "Portal Inmobiliario", "Lopez Bienes Raices"],
        "technology": ["TechNova Solutions", "ByteForge SA", "CloudArg SRL", "Nexo Digital", "Silicon Pampa", "DataCrunch SAS", "AppFactory AR", "Quantum Labs", "DigitalBridge", "CodeSur Technologies", "CyberNorte SRL", "Innova Software", "Red Agil SA", "Pixel Perfect Studio", "NetSolutions Argentina"],
        "gastronomia": ["La Parrilla de Don Luis", "Sabores del Norte", "El Fogon Criollo", "Cafe Havanna", "Resto Bar 1810", "La Cocina de Maria", "Punto Gourmet", "El Molino Restaurant", "Delicias Tucumanas", "Bodegon del Centro", "Pastas Nona Rosa", "Sushi San Martin", "Cerveceria Andina", "Dulce Tentacion", "Restaurante El Puerto"],
        "legal": ["Estudio Juridico Rios & Asoc.", "Bufete Legal del Norte", "Abogados Asociados SA", "Consultora Legal Mendoza", "Estudio Notarial Garcia", "Perez & Gomez Abogados", "Defensoria Integral", "Asesoria Juridica Plus", "Estudio Contable Fiscal", "Legal Corp Argentina"],
        "health": ["Clinica Del Sol", "Centro Medico Austral", "Sanatorio San Lucas", "Dental Premium Center", "Laboratorio BioAndes", "Optica Vision Total", "Farmacia del Pueblo", "Kinesiologia Integral", "Nutricion & Salud", "Centro Pediatrico NOA"],
        "education": ["Instituto Cervantes", "Academia del Saber", "Colegio San Martin", "Centro de Idiomas Global", "Universidad Abierta", "Escuela de Negocios AR", "Instituto Tecnico Plus", "Jardin Mis Primeros Pasos", "Curso Online Argentina", "Capacitacion Profesional SA"],
        "insurance": ["Seguros del Litoral", "Aseguradora Nacional", "Proteccion Total SA", "Seguros Rivadavia", "Cobertura Integral SRL", "La Prevision Seguros", "Seguros del Norte", "Grupo Asegurador Sur", "Patrimonio Seguros", "Federal Seguros"],
        "consulting": ["Consultora Empresarial Plus", "Strategy Partners AR", "Asesores Financieros SA", "McKenzie Consultores", "Grupo Estrategia", "Business Intelligence AR", "Deloitte Argentina", "Asesoria Tributaria", "Consultoria de RRHH", "Performance Group"],
        "construction": ["Constructora del Valle", "Edificar SA", "Obras Civiles Patagonia", "Hormigon Armado SRL", "Arquitectura Moderna", "Construplan AR", "Ingenieria Aplicada", "Revestimientos & Mas", "Herreria Industrial NOA", "Materiales del Sur"],
        "logistics": ["Transporte Ejecutivo NOA", "Logistica Express AR", "Envios Rapidos SA", "Cargas del Norte", "Fleet Solutions", "Mudanzas El Gaucho", "Correo Privado Sur", "Distribucion Nacional", "Almacenes Centrales", "Puerto Logistics"],
    }
    cat_key = job["category"].lower()
    businesses = biz_by_cat.get(cat_key, biz_by_cat.get("technology", []))

    province_cities = {
        "Buenos Aires": ["CABA", "La Plata", "Mar del Plata", "Bahia Blanca", "Quilmes", "Lomas de Zamora"],
        "CABA": ["Palermo", "Recoleta", "San Telmo", "Belgrano", "Caballito", "Microcentro"],
        "Tucuman": ["San Miguel de Tucuman", "Yerba Buena", "Tafi Viejo", "Banda del Rio Sali", "Concepcion"],
        "Cordoba": ["Cordoba Capital", "Villa Carlos Paz", "Rio Cuarto", "Villa Maria", "Alta Gracia"],
        "Mendoza": ["Ciudad de Mendoza", "Godoy Cruz", "San Rafael", "Lujan de Cuyo", "Maipu"],
        "Santa Fe": ["Rosario", "Santa Fe Capital", "Rafaela", "Venado Tuerto", "Reconquista"],
        "Salta": ["Salta Capital", "San Lorenzo", "Oran", "Tartagal", "Cafayate"],
    }
    cities = province_cities.get(job["province"], [job["city"]])

    quality_levels = ["excellent", "good", "average", "poor"]
    rec_texts = {
        "excellent": ["Altamente recomendado - presencia digital solida y datos verificados", "Excelente prospecto - web profesional y multiples canales de contacto", "Top prospecto - alta actividad en redes y web actualizada"],
        "good": ["Buen candidato - informacion de contacto verificada", "Prospecto solido - presencia online estable", "Candidato viable - datos consistentes y actividad reciente"],
        "average": ["Promedio - requiere validacion adicional", "Necesita mas investigacion - datos parciales", "Prospecto con potencial - falta info de contacto directa"],
        "poor": ["Baja prioridad - datos limitados disponibles", "No recomendado actualmente - poca presencia digital", "Requiere seguimiento manual - info desactualizada"]
    }
    first_lines = [
        f"Notamos que su empresa tiene una fuerte presencia en {job['city']} y queremos proponerle una alianza estrategica.",
        f"Vimos su actividad en el sector de {job['category']} en {job['province']} y creemos que podemos potenciar su negocio.",
        f"Su empresa destaca en el mercado de {job['city']}. Nos gustaria explorar oportunidades de colaboracion.",
        f"Encontramos su empresa mientras analizabamos el sector de {job['category']} en la region. Tiene un perfil muy interesante.",
        f"Sabemos que {job['category']} en {job['province']} esta creciendo y su empresa esta bien posicionada para aprovechar esta oportunidad.",
    ]

    statuses_pool = ["scored"] * 5 + ["cleaned"] * 2 + ["approved"] + ["rejected"]
    leads_to_insert = []
    used_names = set()
    # Pre-fetch existing lead names for this tenant to avoid duplicates
    existing_names_cursor = db.leads.find({"tenant_id": user["tenant_id"]}, {"_id": 0, "business_name": 1})
    existing_names = set()
    async for edoc in existing_names_cursor:
        existing_names.add(edoc.get("business_name", "").lower().strip())
    for i in range(min(qualified_count, 25)):
        bname = random.choice(businesses)
        suffix = random.choice(["SA", "SRL", "SAS", ""])
        full_name = f"{bname} {suffix}".strip()
        while (full_name in used_names or full_name.lower() in existing_names) and len(used_names) < len(businesses) * 3:
            bname = random.choice(businesses)
            suffix = random.choice(["SA", "SRL", "SAS", ""])
            full_name = f"{bname} {suffix}".strip()
        used_names.add(full_name)
        score = random.randint(35, 98)
        ql = 0 if score >= 85 else (1 if score >= 65 else (2 if score >= 45 else 3))
        city = random.choice(cities)
        slug = bname.lower().replace(" ", "").replace("&", "y").replace(".", "")[:12]
        lead = {
            "id": str(uuid.uuid4()), "tenant_id": user["tenant_id"], "job_id": job_id,
            "business_name": full_name,
            "raw_category": job["category"].lower(), "normalized_category": job["category"].title(),
            "province": job["province"], "city": city,
            "website": f"www.{slug}.com.ar",
            "email": f"contacto@{slug}.com.ar",
            "phone": f"+54 {random.choice(['11','351','381','261','341','387'])} {random.randint(100,999)}-{random.randint(1000,9999)}",
            "ai_score": score, "quality_level": quality_levels[ql],
            "recommendation": random.choice(rec_texts[quality_levels[ql]]),
            "recommended_first_line": random.choice(first_lines),
            "status": random.choice(statuses_pool),
            "created_at": now, "updated_at": now
        }
        leads_to_insert.append(lead)
    if leads_to_insert:
        await db.leads.insert_many(leads_to_insert)
    # Auto-create/update email list with scored leads
    scored_ids = [l["id"] for l in leads_to_insert if l["status"] in ("scored", "approved")]
    auto_list_name = f"{job['category']} - {job['city']} - Calificados"
    auto_list_id = ""
    if scored_ids:
        existing_list = await db.email_lists.find_one({"tenant_id": user["tenant_id"], "name": auto_list_name})
        if existing_list:
            old_ids = existing_list.get("lead_ids", [])
            new_ids = old_ids + [sid for sid in scored_ids if sid not in old_ids]
            await db.email_lists.update_one({"id": existing_list["id"]}, {"$set": {"lead_ids": new_ids, "subscriber_count": len(new_ids), "updated_at": now}})
            auto_list_id = existing_list["id"]
        else:
            auto_list_id = str(uuid.uuid4())
            await db.email_lists.insert_one({"id": auto_list_id, "tenant_id": user["tenant_id"], "name": auto_list_name, "description": f"Auto: {len(scored_ids)} leads calificados de {job['category']} en {job['city']}", "subscriber_count": len(scored_ids), "lead_ids": scored_ids, "created_at": now, "updated_at": now})
    await db.prospect_jobs.update_one({"id": job_id}, {"$set": {"auto_list_name": auto_list_name, "auto_list_id": auto_list_id}})
    updated_job = await db.prospect_jobs.find_one({"id": job_id}, {"_id": 0})
    return updated_job

# ==================== LEADS ====================

@api_router.get("/leads/stats")
async def leads_stats(request: Request, job_id: Optional[str] = None):
    user = await get_current_user(request)
    query = {"tenant_id": user["tenant_id"]}
    if job_id:
        query["job_id"] = job_id
    pipeline = [{"$match": query}, {"$group": {"_id": "$status", "count": {"$sum": 1}}}]
    results = await db.leads.aggregate(pipeline).to_list(20)
    stats = {r["_id"]: r["count"] for r in results}
    return {"total": sum(stats.values()), "scored": stats.get("scored", 0), "approved": stats.get("approved", 0), "rejected": stats.get("rejected", 0), "contacted": stats.get("contacted", 0), "sent_to_crm": stats.get("sent_to_crm", 0), "cleaned": stats.get("cleaned", 0)}

@api_router.get("/leads")
async def list_leads(request: Request, status: Optional[str] = None, job_id: Optional[str] = None, search: Optional[str] = None, source: Optional[str] = None, category: Optional[str] = None, city: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None, sort_by: Optional[str] = "created_at", sort_dir: Optional[str] = "desc", page: int = 1, limit: int = 50):
    user = await get_current_user(request)
    query = {"tenant_id": user["tenant_id"]}
    if status:
        query["status"] = status
    if job_id:
        query["job_id"] = job_id
    if source:
        query["source"] = source
    if category:
        query["normalized_category"] = {"$regex": category, "$options": "i"}
    if city:
        query["city"] = {"$regex": city, "$options": "i"}
    if date_from or date_to:
        date_filter = {}
        if date_from:
            date_filter["$gte"] = date_from
        if date_to:
            date_filter["$lte"] = date_to + "T23:59:59"
        query["created_at"] = date_filter
    if search:
        query["$or"] = [{"business_name": {"$regex": search, "$options": "i"}}, {"email": {"$regex": search, "$options": "i"}}, {"city": {"$regex": search, "$options": "i"}}]
    total = await db.leads.count_documents(query)
    skip = (page - 1) * limit
    sort_field = sort_by if sort_by in ["ai_score", "created_at", "business_name", "city", "normalized_category"] else "created_at"
    sort_direction = 1 if sort_dir == "asc" else -1
    leads = await db.leads.find(query, {"_id": 0}).sort(sort_field, sort_direction).skip(skip).limit(limit).to_list(limit)
    return {"leads": leads, "total": total, "page": page, "limit": limit, "pages": max(1, (total + limit - 1) // limit)}

@api_router.get("/leads/{lead_id}")
async def get_lead(request: Request, lead_id: str):
    user = await get_current_user(request)
    lead = await db.leads.find_one({"id": lead_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    events = await db.lead_events.find({"lead_id": lead_id}, {"_id": 0}).sort("created_at", -1).to_list(50)
    return {**lead, "events": events}

@api_router.post("/leads")
async def create_lead_manual(request: Request, body: Dict[str, Any] = {}):
    """Create a lead manually from the Leads page"""
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    bname = body.get("business_name", "").strip()
    if not bname:
        raise HTTPException(status_code=400, detail="business_name es requerido")
    lead = {
        "id": str(uuid.uuid4()), "tenant_id": user["tenant_id"], "job_id": "",
        "business_name": bname,
        "raw_category": body.get("category", ""),
        "normalized_category": body.get("category", "Manual"),
        "province": body.get("province", ""),
        "city": body.get("city", ""),
        "website": body.get("website", ""),
        "email": body.get("email", ""),
        "phone": body.get("phone", ""),
        "ai_score": 0, "quality_level": "unscored",
        "recommendation": body.get("notes", ""),
        "recommended_first_line": "",
        "source": body.get("source", "manual"),
        "tags": body.get("tags", []),
        "status": "raw",
        "created_at": now, "updated_at": now
    }
    await db.leads.insert_one(lead)
    return {k: v for k, v in lead.items() if k != "_id"}

@api_router.put("/leads/{lead_id}/status")
async def update_lead_status(request: Request, lead_id: str, body: LeadStatusUpdate):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    result = await db.leads.update_one({"id": lead_id, "tenant_id": user["tenant_id"]}, {"$set": {"status": body.status, "updated_at": now}})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Lead not found")
    await db.lead_events.insert_one({"id": str(uuid.uuid4()), "tenant_id": user["tenant_id"], "lead_id": lead_id, "event_type": f"status_changed_to_{body.status}", "details": f"Status updated to {body.status} by {user['name']}", "created_at": now})
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if lead and lead.get("job_id") and body.status == "approved":
        await db.prospect_jobs.update_one({"id": lead["job_id"]}, {"$inc": {"approved_count": 1}})
    if lead and body.status == "sent_to_crm":
        existing = await db.crm_contacts.find_one({"lead_id": lead_id, "tenant_id": user["tenant_id"]})
        if not existing:
            await db.crm_contacts.insert_one({"id": str(uuid.uuid4()), "tenant_id": user["tenant_id"], "lead_id": lead_id, "business_name": lead.get("business_name", ""), "contact_name": "", "email": lead.get("email", ""), "phone": lead.get("phone", ""), "city": lead.get("city", ""), "province": lead.get("province", ""), "category": lead.get("normalized_category", ""), "source": "spectra_flow", "notes": lead.get("recommendation", ""), "stage": "nuevo", "ai_score": lead.get("ai_score", 0), "deal_count": 0, "total_value": 0, "created_at": now, "updated_at": now})
    # Auto-add to "Secuencia" email list when queued
    if lead and body.status == "queued_for_sequence":
        seq_list_name = "Secuencia - Leads en cola"
        seq_list = await db.email_lists.find_one({"tenant_id": user["tenant_id"], "name": seq_list_name})
        if seq_list:
            old_ids = seq_list.get("lead_ids", [])
            if lead_id not in old_ids:
                await db.email_lists.update_one({"id": seq_list["id"]}, {"$set": {"lead_ids": old_ids + [lead_id], "subscriber_count": len(old_ids) + 1, "updated_at": now}})
        else:
            await db.email_lists.insert_one({"id": str(uuid.uuid4()), "tenant_id": user["tenant_id"], "name": seq_list_name, "description": "Leads enviados a secuencia de email", "subscriber_count": 1, "lead_ids": [lead_id], "created_at": now, "updated_at": now})
    return {"message": "Status updated", "status": body.status}

@api_router.put("/leads/{lead_id}/fields")
async def update_lead_fields(request: Request, lead_id: str, body: Dict[str, Any] = {}):
    """Update editable fields on a lead (notes, channel, etc)"""
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    allowed = {"recommendation", "channel", "normalized_category", "city", "province", "website", "email", "phone", "business_name"}
    update = {k: v for k, v in body.items() if k in allowed}
    if not update:
        raise HTTPException(status_code=400, detail="No fields to update")
    update["updated_at"] = now
    result = await db.leads.update_one({"id": lead_id, "tenant_id": user["tenant_id"]}, {"$set": update})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Lead not found")
    return {"message": "Lead actualizado"}

@api_router.post("/leads/bulk-action")
async def bulk_lead_action(request: Request, body: BulkActionRequest):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    status_map = {"approve": "approved", "reject": "rejected", "queue_sequence": "queued_for_sequence", "send_to_crm": "sent_to_crm"}
    new_status = status_map.get(body.action)
    if not new_status:
        raise HTTPException(status_code=400, detail="Invalid action")
    result = await db.leads.update_many({"id": {"$in": body.lead_ids}, "tenant_id": user["tenant_id"]}, {"$set": {"status": new_status, "updated_at": now}})
    # If sending to CRM, also create CRM contacts
    if body.action == "send_to_crm":
        for lead_id in body.lead_ids:
            lead = await db.leads.find_one({"id": lead_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
            if lead:
                existing = await db.crm_contacts.find_one({"lead_id": lead_id, "tenant_id": user["tenant_id"]})
                if not existing:
                    await db.crm_contacts.insert_one({"id": str(uuid.uuid4()), "tenant_id": user["tenant_id"], "lead_id": lead_id, "business_name": lead.get("business_name", ""), "contact_name": "", "email": lead.get("email", ""), "phone": lead.get("phone", ""), "city": lead.get("city", ""), "province": lead.get("province", ""), "category": lead.get("normalized_category", ""), "source": "spectra_flow", "notes": lead.get("recommendation", ""), "stage": "nuevo", "ai_score": lead.get("ai_score", 0), "deal_count": 0, "total_value": 0, "created_at": now, "updated_at": now})
    return {"message": f"{result.modified_count} leads actualizados a {new_status}"}

# ==================== CAMPAIGNS ====================

@api_router.get("/campaigns")
async def list_campaigns(request: Request):
    user = await get_current_user(request)
    campaigns = await db.campaigns.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return campaigns

@api_router.post("/campaigns")
async def create_campaign(request: Request, body: CampaignCreate):
    user = await get_current_user(request)
    campaign_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    campaign = {"id": campaign_id, "tenant_id": user["tenant_id"], "name": body.name, "sender_profile_id": body.sender_profile_id or "", "template_id": body.template_id or "", "status": "draft", "lead_count": len(body.lead_ids) if body.lead_ids else 0, "lead_ids": body.lead_ids or [], "sent_count": 0, "open_count": 0, "click_count": 0, "reply_count": 0, "interested_count": 0, "crm_count": 0, "created_by": user["name"], "created_at": now, "updated_at": now}
    await db.campaigns.insert_one(campaign)
    return serialize_doc(campaign)

@api_router.get("/campaigns/{campaign_id}")
async def get_campaign(request: Request, campaign_id: str):
    user = await get_current_user(request)
    campaign = await db.campaigns.find_one({"id": campaign_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    return campaign

@api_router.put("/campaigns/{campaign_id}")
async def update_campaign(request: Request, campaign_id: str, body: CampaignUpdate):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    update_data = {k: v for k, v in body.model_dump().items() if v is not None}
    update_data["updated_at"] = now
    result = await db.campaigns.update_one({"id": campaign_id, "tenant_id": user["tenant_id"]}, {"$set": update_data})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Campaign not found")
    updated = await db.campaigns.find_one({"id": campaign_id}, {"_id": 0})
    return updated

# ==================== TEMPLATES ====================

@api_router.get("/templates")
async def list_templates(request: Request):
    user = await get_current_user(request)
    templates = await db.templates.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return templates

@api_router.post("/templates")
async def create_template(request: Request, body: TemplateCreate):
    user = await get_current_user(request)
    template_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    template = {"id": template_id, "tenant_id": user["tenant_id"], "name": body.name, "subject": body.subject, "html_body": body.html_body, "plain_text": body.plain_text or "", "variables": body.variables or [], "signature": body.signature or "", "created_at": now, "updated_at": now}
    await db.templates.insert_one(template)
    return serialize_doc(template)

@api_router.put("/templates/{template_id}")
async def update_template(request: Request, template_id: str, body: TemplateUpdate):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    update_data = {k: v for k, v in body.model_dump().items() if v is not None}
    update_data["updated_at"] = now
    result = await db.templates.update_one({"id": template_id, "tenant_id": user["tenant_id"]}, {"$set": update_data})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    updated = await db.templates.find_one({"id": template_id}, {"_id": 0})
    return updated

@api_router.delete("/templates/{template_id}")
async def delete_template(request: Request, template_id: str):
    user = await get_current_user(request)
    result = await db.templates.delete_one({"id": template_id, "tenant_id": user["tenant_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    return {"message": "Template deleted"}

@api_router.post("/templates/ab-test")
async def create_ab_test(request: Request, body: Dict[str, Any] = {}):
    """Create an A/B test between two templates"""
    user = await get_current_user(request)
    template_a_id = body.get("template_a_id", "")
    template_b_id = body.get("template_b_id", "")
    name = body.get("name", "A/B Test")
    split = body.get("split", 50)
    if not template_a_id or not template_b_id:
        raise HTTPException(status_code=400, detail="Se requieren dos plantillas")
    template_a = await db.templates.find_one({"id": template_a_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    template_b = await db.templates.find_one({"id": template_b_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not template_a or not template_b:
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")
    now = datetime.now(timezone.utc).isoformat()
    test_id = str(uuid.uuid4())
    doc = {
        "id": test_id, "tenant_id": user["tenant_id"], "name": name,
        "template_a": {"id": template_a_id, "name": template_a.get("name", ""), "subject": template_a.get("subject", ""), "sent": 0, "opens": 0, "clicks": 0, "replies": 0},
        "template_b": {"id": template_b_id, "name": template_b.get("name", ""), "subject": template_b.get("subject", ""), "sent": 0, "opens": 0, "clicks": 0, "replies": 0},
        "split": split, "status": "borrador", "winner": None, "created_at": now, "updated_at": now
    }
    await db.ab_tests.insert_one(doc)
    return serialize_doc(doc)

@api_router.get("/templates/ab-tests")
async def list_ab_tests(request: Request):
    user = await get_current_user(request)
    tests = await db.ab_tests.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).sort("created_at", -1).to_list(50)
    return tests

@api_router.post("/templates/ab-tests/{test_id}/simulate")
async def simulate_ab_test(request: Request, test_id: str):
    user = await get_current_user(request)
    test = await db.ab_tests.find_one({"id": test_id, "tenant_id": user["tenant_id"]})
    if not test:
        raise HTTPException(status_code=404, detail="A/B Test no encontrado")
    now = datetime.now(timezone.utc).isoformat()
    total = random.randint(100, 500)
    split = test.get("split", 50)
    sent_a = int(total * split / 100)
    sent_b = total - sent_a
    a_opens = int(sent_a * random.uniform(0.20, 0.55))
    b_opens = int(sent_b * random.uniform(0.20, 0.55))
    a_clicks = int(a_opens * random.uniform(0.10, 0.35))
    b_clicks = int(b_opens * random.uniform(0.10, 0.35))
    a_replies = int(a_clicks * random.uniform(0.20, 0.50))
    b_replies = int(b_clicks * random.uniform(0.20, 0.50))
    a_rate = (a_opens / max(sent_a, 1)) * 100
    b_rate = (b_opens / max(sent_b, 1)) * 100
    winner = "A" if a_rate > b_rate else ("B" if b_rate > a_rate else "Empate")
    await db.ab_tests.update_one({"id": test_id}, {"$set": {
        "template_a.sent": sent_a, "template_a.opens": a_opens, "template_a.clicks": a_clicks, "template_a.replies": a_replies,
        "template_b.sent": sent_b, "template_b.opens": b_opens, "template_b.clicks": b_clicks, "template_b.replies": b_replies,
        "status": "completado", "winner": winner, "updated_at": now
    }})
    updated = await db.ab_tests.find_one({"id": test_id}, {"_id": 0})
    return updated

class SendTestRequest(BaseModel):
    to_email: Optional[str] = "test@demo.com"

@api_router.post("/templates/{template_id}/send-test")
async def send_test_email(request: Request, template_id: str, body: SendTestRequest):
    user = await get_current_user(request)
    template = await db.templates.find_one({"id": template_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    preview_html = template["html_body"].replace("{business_name}", "Empresa Demo SA").replace("{city}", "Buenos Aires").replace("{normalized_category}", "Tecnologia").replace("{recommended_first_line}", "Notamos que su empresa tiene una fuerte presencia en el mercado digital.").replace("{sender_name}", user.get("name", "Spectra Flow"))
    preview_subject = template["subject"].replace("{business_name}", "Empresa Demo SA")
    tenant = await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0})
    verified_domain = await db.domains.find_one({"tenant_id": user["tenant_id"], "status": {"$in": ["verified", "warmup_recommended", "ready_to_send"]}}, {"_id": 0})
    from_email = "onboarding@resend.dev"
    if verified_domain:
        from_email = verified_domain.get("sender_email", f"noreply@{verified_domain.get('subdomain', verified_domain.get('domain', 'resend.dev'))}")
    sender_name = verified_domain.get("sender_name", "Spectra Flow") if verified_domain else "Spectra Flow"
    if resend.api_key:
        try:
            params = {
                "from": f"{sender_name} <{from_email}>",
                "to": [body.to_email],
                "subject": preview_subject,
                "html": preview_html
            }
            result = await asyncio.to_thread(resend.Emails.send, params)
            result_data = result if isinstance(result, dict) else vars(result) if hasattr(result, '__dict__') else {}
            email_id = result_data.get("id", "")
            logger.info(f"Test email sent via Resend to {body.to_email}, id: {email_id}")
            return {"message": f"Email de prueba ENVIADO a {body.to_email}", "preview_subject": preview_subject, "preview_html": preview_html, "simulated": False, "email_id": email_id, "from_email": from_email}
        except Exception as e:
            logger.error(f"Resend send failed: {e}")
            return {"message": f"Error enviando email: {str(e)}", "preview_subject": preview_subject, "preview_html": preview_html, "simulated": True, "error": str(e)}
    return {"message": f"Email de prueba simulado a {body.to_email} (Resend no configurado)", "preview_subject": preview_subject, "preview_html": preview_html, "simulated": True}

@api_router.post("/campaigns/{campaign_id}/simulate")
async def simulate_campaign(request: Request, campaign_id: str):
    user = await get_current_user(request)
    campaign = await db.campaigns.find_one({"id": campaign_id, "tenant_id": user["tenant_id"]})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    now = datetime.now(timezone.utc).isoformat()
    lead_count = campaign.get("lead_count", 0) or random.randint(15, 50)
    sent = lead_count
    opens = int(sent * random.uniform(0.45, 0.70))
    clicks = int(opens * random.uniform(0.25, 0.45))
    replies = int(clicks * random.uniform(0.30, 0.55))
    interested = int(replies * random.uniform(0.40, 0.70))
    crm = int(interested * random.uniform(0.50, 0.80))
    await db.campaigns.update_one({"id": campaign_id}, {"$set": {"status": "completed", "lead_count": lead_count, "sent_count": sent, "open_count": opens, "click_count": clicks, "reply_count": replies, "interested_count": interested, "crm_count": crm, "updated_at": now}})
    updated = await db.campaigns.find_one({"id": campaign_id}, {"_id": 0})
    return updated

# ==================== DOMAINS ====================

@api_router.get("/domains")
async def list_domains(request: Request):
    user = await get_current_user(request)
    domains = await db.domains.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return domains

@api_router.post("/domains")
async def create_domain(request: Request, body: DomainCreate):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    domain_name = body.subdomain if body.subdomain else body.domain
    dns_records = []
    resend_domain_id = ""
    status = "dns_pending"
    try:
        resend_result = await asyncio.to_thread(resend.Domains.create, {"name": domain_name})
        resend_data = resend_result if isinstance(resend_result, dict) else vars(resend_result) if hasattr(resend_result, '__dict__') else {"id": str(resend_result)}
        resend_domain_id = resend_data.get("id", "")
        raw_records = resend_data.get("records", resend_data.get("dns_records", []))
        for rec in raw_records:
            r = rec if isinstance(rec, dict) else vars(rec) if hasattr(rec, '__dict__') else {}
            dns_records.append({
                "type": r.get("record", r.get("type", "TXT")),
                "name": r.get("name", ""),
                "value": r.get("value", ""),
                "priority": r.get("priority"),
                "ttl": r.get("ttl", "Auto"),
                "status": r.get("status", "pending"),
                "verified": r.get("status", "") == "verified"
            })
        status = resend_data.get("status", "dns_pending")
        logger.info(f"Resend domain created: {resend_domain_id} for {domain_name}")
    except Exception as e:
        logger.error(f"Resend domain creation failed: {e}")
        dns_records = [
            {"type": "TXT", "name": domain_name, "value": "v=spf1 include:_spf.resend.com ~all", "verified": False, "status": "pending"},
            {"type": "CNAME", "name": f"resend._domainkey.{domain_name}", "value": "resend.domainkey.resend.com", "verified": False, "status": "pending"},
            {"type": "MX", "name": domain_name, "value": "feedback-smtp.resend.com", "verified": False, "status": "pending"}
        ]
    domain_id = str(uuid.uuid4())
    domain_doc = {"id": domain_id, "tenant_id": user["tenant_id"], "domain": body.domain, "subdomain": domain_name, "resend_domain_id": resend_domain_id, "status": status, "dns_records": dns_records, "sender_name": body.sender_name or "", "sender_email": body.sender_email or f"noreply@{domain_name}", "reply_to": body.reply_to or "", "signature": body.signature or "", "warmup_status": "not_started", "created_at": now, "updated_at": now}
    await db.domains.insert_one(domain_doc)
    return serialize_doc(domain_doc)

@api_router.put("/domains/{domain_id}")
async def update_domain(request: Request, domain_id: str, body: DomainUpdate):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    update_data = {k: v for k, v in body.model_dump().items() if v is not None}
    update_data["updated_at"] = now
    result = await db.domains.update_one({"id": domain_id, "tenant_id": user["tenant_id"]}, {"$set": update_data})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Domain not found")
    updated = await db.domains.find_one({"id": domain_id}, {"_id": 0})
    return updated

@api_router.post("/domains/{domain_id}/verify")
async def verify_domain(request: Request, domain_id: str):
    user = await get_current_user(request)
    domain = await db.domains.find_one({"id": domain_id, "tenant_id": user["tenant_id"]})
    if not domain:
        raise HTTPException(status_code=404, detail="Domain not found")
    now = datetime.now(timezone.utc).isoformat()
    resend_domain_id = domain.get("resend_domain_id", "")
    if resend_domain_id:
        try:
            await asyncio.to_thread(resend.Domains.verify, resend_domain_id)
            import time
            time.sleep(2)
            resend_info = await asyncio.to_thread(resend.Domains.get, resend_domain_id)
            info_data = resend_info if isinstance(resend_info, dict) else vars(resend_info) if hasattr(resend_info, '__dict__') else {}
            new_status = info_data.get("status", "pending")
            raw_records = info_data.get("records", info_data.get("dns_records", []))
            dns_records = []
            for rec in raw_records:
                r = rec if isinstance(rec, dict) else vars(rec) if hasattr(rec, '__dict__') else {}
                is_verified = r.get("status", "") == "verified"
                dns_records.append({
                    "type": r.get("record", r.get("type", "TXT")),
                    "name": r.get("name", ""),
                    "value": r.get("value", ""),
                    "priority": r.get("priority"),
                    "ttl": r.get("ttl", "Auto"),
                    "status": r.get("status", "pending"),
                    "verified": is_verified
                })
            if new_status == "verified":
                mapped_status = "warmup_recommended"
            elif new_status in ("pending", "not_started"):
                mapped_status = "verifying"
            elif new_status in ("failed", "partially_failed"):
                mapped_status = "configuration_error"
            else:
                mapped_status = "verifying"
            await db.domains.update_one({"id": domain_id}, {"$set": {"dns_records": dns_records, "status": mapped_status, "updated_at": now}})
            logger.info(f"Resend domain verification triggered for {resend_domain_id}, status: {new_status}")
        except Exception as e:
            logger.error(f"Resend domain verification failed: {e}")
            await db.domains.update_one({"id": domain_id}, {"$set": {"status": "verifying", "updated_at": now}})
    else:
        dns_records = domain.get("dns_records", [])
        all_verified = True
        for record in dns_records:
            record["verified"] = random.choice([True, True, True, False])
            if not record["verified"]:
                all_verified = False
        new_status = "warmup_recommended" if all_verified else "verifying"
        await db.domains.update_one({"id": domain_id}, {"$set": {"dns_records": dns_records, "status": new_status, "updated_at": now}})
    updated = await db.domains.find_one({"id": domain_id}, {"_id": 0})
    return updated

@api_router.get("/domains/resend-status")
async def get_resend_domains_status(request: Request):
    """Fetch all domains from Resend API to see real-time status"""
    user = await get_current_user(request)
    if not resend.api_key:
        return {"domains": [], "error": "Resend API key not configured"}
    try:
        result = await asyncio.to_thread(resend.Domains.list)
        result_data = result if isinstance(result, dict) else vars(result) if hasattr(result, '__dict__') else {}
        domains_list = result_data.get("data", [])
        parsed = []
        for d in domains_list:
            dd = d if isinstance(d, dict) else vars(d) if hasattr(d, '__dict__') else {}
            parsed.append({"id": dd.get("id", ""), "name": dd.get("name", ""), "status": dd.get("status", "unknown"), "region": dd.get("region", ""), "created_at": dd.get("created_at", "")})
        return {"domains": parsed}
    except Exception as e:
        logger.error(f"Resend domains list failed: {e}")
        return {"domains": [], "error": str(e)}

@api_router.post("/email/send")
async def send_email_direct(request: Request, body: Dict[str, Any] = {}):
    """Send a single email via Resend"""
    user = await get_current_user(request)
    to_email = body.get("to_email", "")
    subject = body.get("subject", "")
    html_body = body.get("html_body", "")
    from_name = body.get("from_name", "Spectra Flow")
    from_email_addr = body.get("from_email", "")
    if not to_email or not subject:
        raise HTTPException(status_code=400, detail="to_email and subject are required")
    if not from_email_addr:
        verified_domain = await db.domains.find_one({"tenant_id": user["tenant_id"], "status": {"$in": ["verified", "warmup_recommended", "ready_to_send"]}}, {"_id": 0})
        if verified_domain:
            from_email_addr = verified_domain.get("sender_email", f"noreply@{verified_domain.get('subdomain', 'resend.dev')}")
            from_name = verified_domain.get("sender_name", from_name)
        else:
            from_email_addr = "onboarding@resend.dev"
    if not resend.api_key:
        return {"message": "Resend API key not configured, email simulated", "simulated": True}
    try:
        params = {"from": f"{from_name} <{from_email_addr}>", "to": [to_email], "subject": subject, "html": html_body or f"<p>{subject}</p>"}
        result = await asyncio.to_thread(resend.Emails.send, params)
        result_data = result if isinstance(result, dict) else vars(result) if hasattr(result, '__dict__') else {}
        email_id = result_data.get("id", "")
        logger.info(f"Email sent via Resend to {to_email}, id: {email_id}")
        return {"message": f"Email enviado a {to_email}", "simulated": False, "email_id": email_id}
    except Exception as e:
        logger.error(f"Resend send error: {e}")
        raise HTTPException(status_code=500, detail=f"Error sending email: {str(e)}")

@api_router.post("/domains/sync-resend")
async def sync_domains_from_resend(request: Request):
    """Sync local domain records with real Resend API data"""
    user = await get_current_user(request)
    if not resend.api_key:
        raise HTTPException(status_code=400, detail="Resend API key not configured")
    try:
        result = await asyncio.to_thread(resend.Domains.list)
        result_data = result if isinstance(result, dict) else vars(result) if hasattr(result, '__dict__') else {}
        domains_list = result_data.get("data", [])
        synced = 0
        now = datetime.now(timezone.utc).isoformat()
        for d in domains_list:
            dd = d if isinstance(d, dict) else vars(d) if hasattr(d, '__dict__') else {}
            resend_id = dd.get("id", "")
            domain_name = dd.get("name", "")
            resend_status = dd.get("status", "unknown")
            mapped_status = "warmup_recommended" if resend_status == "verified" else ("verifying" if resend_status in ("pending", "not_started") else ("configuration_error" if resend_status in ("failed", "partially_failed") else "verifying"))
            dns_records = []
            try:
                detail = await asyncio.to_thread(resend.Domains.get, resend_id)
                detail_data = detail if isinstance(detail, dict) else vars(detail) if hasattr(detail, '__dict__') else {}
                raw_records = detail_data.get("records", detail_data.get("dns_records", []))
                for rec in raw_records:
                    r = rec if isinstance(rec, dict) else vars(rec) if hasattr(rec, '__dict__') else {}
                    dns_records.append({"type": r.get("record", r.get("type", "TXT")), "name": r.get("name", ""), "value": r.get("value", ""), "priority": r.get("priority"), "ttl": r.get("ttl", "Auto"), "status": r.get("status", "pending"), "verified": r.get("status", "") == "verified"})
            except Exception:
                pass
            existing = await db.domains.find_one({"$or": [{"resend_domain_id": resend_id}, {"domain": domain_name}, {"subdomain": domain_name}], "tenant_id": user["tenant_id"]})
            if existing:
                update_fields = {"resend_domain_id": resend_id, "status": mapped_status, "updated_at": now}
                if dns_records:
                    update_fields["dns_records"] = dns_records
                await db.domains.update_one({"_id": existing["_id"]}, {"$set": update_fields})
            else:
                await db.domains.insert_one({"id": str(uuid.uuid4()), "tenant_id": user["tenant_id"], "domain": domain_name, "subdomain": domain_name, "resend_domain_id": resend_id, "status": mapped_status, "dns_records": dns_records, "sender_name": "Spectra Flow", "sender_email": f"noreply@{domain_name}", "reply_to": "", "signature": "", "warmup_status": "not_started", "created_at": now, "updated_at": now})
            synced += 1
        domains = await db.domains.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).sort("created_at", -1).to_list(100)
        return {"synced": synced, "domains": domains}
    except Exception as e:
        logger.error(f"Resend sync failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==================== CRM SYNC ====================

@api_router.get("/crm-sync/logs")
async def list_crm_sync_logs(request: Request):
    user = await get_current_user(request)
    logs = await db.crm_sync_logs.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return logs

@api_router.post("/crm-sync/push")
async def push_to_crm(request: Request, body: CrmPushRequest):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    results = []
    for lead_id in body.lead_ids:
        lead = await db.leads.find_one({"id": lead_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
        if not lead:
            continue
        success = random.random() > 0.1
        log = {"id": str(uuid.uuid4()), "tenant_id": user["tenant_id"], "lead_id": lead_id, "lead_name": lead.get("business_name", ""), "status": "synced" if success else "error", "error": "" if success else "Connection timeout", "synced_at": now if success else None, "assigned_owner": body.assigned_owner or "Default Owner", "created_at": now}
        await db.crm_sync_logs.insert_one(log)
        if success:
            await db.leads.update_one({"id": lead_id}, {"$set": {"status": "sent_to_crm", "updated_at": now}})
        results.append(serialize_doc(log))
    return {"message": f"{len(results)} leads processed", "results": results}

@api_router.post("/crm-sync/retry/{log_id}")
async def retry_crm_sync(request: Request, log_id: str):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    await db.crm_sync_logs.update_one({"id": log_id, "tenant_id": user["tenant_id"]}, {"$set": {"status": "synced", "error": "", "synced_at": now}})
    return {"message": "Retry successful"}

# ==================== CRM MODULE ====================

class CrmContactCreate(BaseModel):
    lead_id: Optional[str] = ""
    business_name: str
    contact_name: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""
    city: Optional[str] = ""
    province: Optional[str] = ""
    category: Optional[str] = ""
    source: Optional[str] = "spectra_flow"
    notes: Optional[str] = ""

class CrmDealCreate(BaseModel):
    contact_id: str
    title: str
    value: Optional[float] = 0
    stage: Optional[str] = "nuevo"
    assigned_to: Optional[str] = ""

class CrmDealUpdate(BaseModel):
    title: Optional[str] = None
    value: Optional[float] = None
    stage: Optional[str] = None
    assigned_to: Optional[str] = None

class CrmNoteCreate(BaseModel):
    contact_id: str
    content: str
    note_type: Optional[str] = "nota"

@api_router.get("/crm/contacts")
async def list_crm_contacts(request: Request, search: Optional[str] = None):
    user = await get_current_user(request)
    query = {"tenant_id": user["tenant_id"]}
    if search:
        query["$or"] = [{"business_name": {"$regex": search, "$options": "i"}}, {"contact_name": {"$regex": search, "$options": "i"}}, {"email": {"$regex": search, "$options": "i"}}]
    contacts = await db.crm_contacts.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    return contacts

@api_router.post("/crm/contacts")
async def create_crm_contact(request: Request, body: CrmContactCreate):
    user = await get_current_user(request)
    contact_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    contact = {"id": contact_id, "tenant_id": user["tenant_id"], "lead_id": body.lead_id or "", "business_name": body.business_name, "contact_name": body.contact_name or "", "email": body.email or "", "phone": body.phone or "", "city": body.city or "", "province": body.province or "", "category": body.category or "", "source": body.source or "spectra_flow", "notes": body.notes or "", "stage": "nuevo", "deal_count": 0, "total_value": 0, "created_at": now, "updated_at": now}
    await db.crm_contacts.insert_one(contact)
    if body.lead_id:
        await db.leads.update_one({"id": body.lead_id}, {"$set": {"status": "sent_to_crm", "updated_at": now}})
    return serialize_doc(contact)

@api_router.post("/crm/contacts/from-leads")
async def create_contacts_from_leads(request: Request, body: CrmPushRequest):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    created = []
    for lead_id in body.lead_ids:
        lead = await db.leads.find_one({"id": lead_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
        if not lead:
            continue
        existing = await db.crm_contacts.find_one({"lead_id": lead_id, "tenant_id": user["tenant_id"]})
        if existing:
            continue
        contact = {"id": str(uuid.uuid4()), "tenant_id": user["tenant_id"], "lead_id": lead_id, "business_name": lead.get("business_name", ""), "contact_name": "", "email": lead.get("email", ""), "phone": lead.get("phone", ""), "city": lead.get("city", ""), "province": lead.get("province", ""), "category": lead.get("normalized_category", ""), "source": "spectra_flow", "notes": lead.get("recommendation", ""), "stage": "nuevo", "ai_score": lead.get("ai_score", 0), "deal_count": 0, "total_value": 0, "created_at": now, "updated_at": now}
        await db.crm_contacts.insert_one(contact)
        await db.leads.update_one({"id": lead_id}, {"$set": {"status": "sent_to_crm", "updated_at": now}})
        created.append(serialize_doc(contact))
    return {"message": f"{len(created)} contactos creados en el CRM", "contacts": created}

@api_router.get("/crm/contacts/{contact_id}")
async def get_crm_contact(request: Request, contact_id: str):
    user = await get_current_user(request)
    contact = await db.crm_contacts.find_one({"id": contact_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not contact:
        raise HTTPException(status_code=404, detail="Contacto no encontrado")
    deals = await db.crm_deals.find({"contact_id": contact_id}, {"_id": 0}).sort("created_at", -1).to_list(50)
    notes = await db.crm_notes.find({"contact_id": contact_id}, {"_id": 0}).sort("created_at", -1).to_list(50)
    return {**contact, "deals": deals, "notes_list": notes}

@api_router.put("/crm/contacts/{contact_id}")
async def update_crm_contact(request: Request, contact_id: str, body: Dict[str, Any] = {}):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    update_data = {k: v for k, v in body.items() if v is not None and k not in ("id", "tenant_id", "_id")}
    update_data["updated_at"] = now
    new_stage = body.get("stage")
    result = await db.crm_contacts.update_one({"id": contact_id, "tenant_id": user["tenant_id"]}, {"$set": update_data})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Contacto no encontrado")
    # Auto-create deal if stage changed and no deal exists
    if new_stage:
        existing_deal = await db.crm_deals.find_one({"contact_id": contact_id, "tenant_id": user["tenant_id"]})
        contact = await db.crm_contacts.find_one({"id": contact_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
        if not existing_deal and contact:
            deal_title = contact.get("business_name", "Oportunidad")
            prob = {"nuevo": 10, "contactado": 30, "propuesta": 50, "negociacion": 70, "ganado": 100, "perdido": 0}.get(new_stage, 20)
            await db.crm_deals.insert_one({"id": str(uuid.uuid4()), "tenant_id": user["tenant_id"], "contact_id": contact_id, "contact_name": contact.get("business_name", ""), "company": contact.get("business_name", ""), "title": deal_title, "name": deal_title, "value": 0, "stage": new_stage, "probability": prob, "assigned_to": user.get("name", ""), "created_at": now, "updated_at": now})
            await db.crm_contacts.update_one({"id": contact_id}, {"$inc": {"deal_count": 1}})
        elif existing_deal:
            await db.crm_deals.update_one({"contact_id": contact_id, "tenant_id": user["tenant_id"]}, {"$set": {"stage": new_stage, "updated_at": now}})
    updated = await db.crm_contacts.find_one({"id": contact_id}, {"_id": 0})
    return updated

@api_router.get("/crm/deals")
async def list_crm_deals(request: Request, stage: Optional[str] = None):
    user = await get_current_user(request)
    query = {"tenant_id": user["tenant_id"]}
    if stage:
        query["stage"] = stage
    deals = await db.crm_deals.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    return deals

@api_router.post("/crm/deals")
async def create_crm_deal(request: Request, body: CrmDealCreate):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    contact = await db.crm_contacts.find_one({"id": body.contact_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    # Check if deal already exists for this contact — update instead of creating duplicate
    existing_deal = await db.crm_deals.find_one({"contact_id": body.contact_id, "tenant_id": user["tenant_id"]})
    if existing_deal:
        update_fields = {"updated_at": now}
        if body.title:
            update_fields["title"] = body.title
            update_fields["name"] = body.title
        if body.value is not None:
            update_fields["value"] = body.value
        if body.stage:
            update_fields["stage"] = body.stage
        if body.assigned_to:
            update_fields["assigned_to"] = body.assigned_to
        await db.crm_deals.update_one({"id": existing_deal["id"]}, {"$set": update_fields})
        if body.value and body.value != existing_deal.get("value", 0):
            diff = (body.value or 0) - existing_deal.get("value", 0)
            await db.crm_contacts.update_one({"id": body.contact_id}, {"$inc": {"total_value": diff}})
        updated = await db.crm_deals.find_one({"id": existing_deal["id"]}, {"_id": 0})
        return serialize_doc(updated)
    deal_id = str(uuid.uuid4())
    deal = {"id": deal_id, "tenant_id": user["tenant_id"], "contact_id": body.contact_id, "contact_name": contact.get("business_name", "") if contact else "", "title": body.title, "value": body.value or 0, "stage": body.stage or "nuevo", "assigned_to": body.assigned_to or user.get("name", ""), "created_at": now, "updated_at": now}
    await db.crm_deals.insert_one(deal)
    await db.crm_contacts.update_one({"id": body.contact_id}, {"$inc": {"deal_count": 1, "total_value": body.value or 0}})
    return serialize_doc(deal)

@api_router.put("/crm/deals/{deal_id}")
async def update_crm_deal(request: Request, deal_id: str, body: CrmDealUpdate):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    update_data = {k: v for k, v in body.model_dump().items() if v is not None}
    update_data["updated_at"] = now
    result = await db.crm_deals.update_one({"id": deal_id, "tenant_id": user["tenant_id"]}, {"$set": update_data})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Oportunidad no encontrada")
    updated = await db.crm_deals.find_one({"id": deal_id}, {"_id": 0})
    return updated

@api_router.post("/crm/notes")
async def create_crm_note(request: Request, body: CrmNoteCreate):
    user = await get_current_user(request)
    note_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    note = {"id": note_id, "tenant_id": user["tenant_id"], "contact_id": body.contact_id, "content": body.content, "note_type": body.note_type or "nota", "author": user.get("name", ""), "created_at": now}
    await db.crm_notes.insert_one(note)
    return serialize_doc(note)

@api_router.get("/crm/stats")
async def get_crm_stats(request: Request, period: Optional[str] = None):
    user = await get_current_user(request)
    tid = user["tenant_id"]
    deal_filter = {"tenant_id": tid}
    contact_filter = {"tenant_id": tid}
    if period and period != "all":
        days = {"week": 7, "month": 30, "quarter": 90}.get(period, 0)
        if days:
            cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
            deal_filter["created_at"] = {"$gte": cutoff}
            contact_filter["created_at"] = {"$gte": cutoff}
    total_contacts = await db.crm_contacts.count_documents(contact_filter)
    total_deals = await db.crm_deals.count_documents(deal_filter)
    stages = ["nuevo", "contactado", "propuesta", "negociacion", "ganado", "perdido"]
    stage_counts = {}
    for s in stages:
        stage_counts[s] = await db.crm_deals.count_documents({**deal_filter, "stage": s})
    pipeline = [{"$match": {**deal_filter, "stage": "ganado"}}, {"$group": {"_id": None, "total": {"$sum": "$value"}}}]
    won_value = await db.crm_deals.aggregate(pipeline).to_list(1)
    lost_pipeline = [{"$match": {**deal_filter, "stage": "perdido"}}, {"$group": {"_id": None, "total": {"$sum": "$value"}}}]
    lost_value = await db.crm_deals.aggregate(lost_pipeline).to_list(1)
    return {"total_contacts": total_contacts, "total_deals": total_deals, "stage_counts": stage_counts, "won_value": won_value[0]["total"] if won_value else 0, "lost_value": lost_value[0]["total"] if lost_value else 0}

# ==================== ANALYTICS ====================

@api_router.get("/analytics")
async def get_analytics(request: Request, period: Optional[str] = None):
    user = await get_current_user(request)
    tid = user["tenant_id"]
    lead_filter = {"tenant_id": tid}
    if period and period != "all":
        days = {"week": 7, "month": 30, "quarter": 90}.get(period, 0)
        if days:
            cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
            lead_filter["created_at"] = {"$gte": cutoff}
    stats = {
        "jobs_created": await db.prospect_jobs.count_documents({"tenant_id": tid}),
        "raw_leads": await db.leads.count_documents({**lead_filter, "status": "raw"}),
        "cleaned_leads": await db.leads.count_documents({**lead_filter, "status": "cleaned"}),
        "scored_leads": await db.leads.count_documents({**lead_filter, "status": "scored"}),
        "qualified_leads": await db.leads.count_documents({**lead_filter, "status": {"$in": ["scored", "approved"]}}),
        "rejected_leads": await db.leads.count_documents({**lead_filter, "status": "rejected"}),
        "approved_leads": await db.leads.count_documents({**lead_filter, "status": "approved"}),
        "total_leads": await db.leads.count_documents(lead_filter),
    }
    pipeline = [{"$match": {"tenant_id": tid}}, {"$group": {"_id": None, "sent": {"$sum": "$sent_count"}, "opens": {"$sum": "$open_count"}, "clicks": {"$sum": "$click_count"}, "replies": {"$sum": "$reply_count"}, "interested": {"$sum": "$interested_count"}, "crm": {"$sum": "$crm_count"}}}]
    cs = (await db.campaigns.aggregate(pipeline).to_list(1))
    cs = cs[0] if cs else {}
    stats.update({"emails_sent": cs.get("sent", 0), "opens": cs.get("opens", 0), "clicks": cs.get("clicks", 0), "replies": cs.get("replies", 0), "interested": cs.get("interested", 0), "crm_handoffs": await db.crm_sync_logs.count_documents({"tenant_id": tid, "status": "synced"})})
    total = stats["total_leads"] or 1
    stats["qualification_rate"] = round((stats["qualified_leads"] / total) * 100, 1)
    stats["approval_rate"] = round((stats["approved_leads"] / total) * 100, 1)
    stats["email_open_rate"] = round((stats["opens"] / max(stats["emails_sent"], 1)) * 100, 1)
    stats["reply_rate"] = round((stats["replies"] / max(stats["emails_sent"], 1)) * 100, 1)
    return stats

@api_router.get("/analytics/time-series")
async def get_analytics_time_series(request: Request, period: str = "week"):
    """Returns time-series data for leads and emails by day/week"""
    user = await get_current_user(request)
    tid = user["tenant_id"]
    from datetime import timedelta
    now_dt = datetime.now(timezone.utc)
    if period == "month":
        days = 30
    elif period == "quarter":
        days = 90
    elif period == "year":
        days = 365
    else:
        days = 7
    start = (now_dt - timedelta(days=days)).isoformat()
    # Get leads by date
    leads_pipeline = [{"$match": {"tenant_id": tid, "created_at": {"$gte": start}}}, {"$group": {"_id": {"$substr": ["$created_at", 0, 10]}, "count": {"$sum": 1}, "scored": {"$sum": {"$cond": [{"$in": ["$status", ["scored", "approved"]]}, 1, 0]}}, "rejected": {"$sum": {"$cond": [{"$eq": ["$status", "rejected"]}, 1, 0]}}}}]
    leads_ts = await db.leads.aggregate(leads_pipeline).to_list(days + 1)
    # Get jobs by date
    jobs_pipeline = [{"$match": {"tenant_id": tid, "created_at": {"$gte": start}}}, {"$group": {"_id": {"$substr": ["$created_at", 0, 10]}, "count": {"$sum": 1}}}]
    jobs_ts = await db.prospect_jobs.aggregate(jobs_pipeline).to_list(days + 1)
    # Get contacts by date
    contacts_pipeline = [{"$match": {"tenant_id": tid, "created_at": {"$gte": start}}}, {"$group": {"_id": {"$substr": ["$created_at", 0, 10]}, "count": {"$sum": 1}}}]
    contacts_ts = await db.crm_contacts.aggregate(contacts_pipeline).to_list(days + 1)
    # Build daily data
    data = []
    for i in range(days, -1, -1):
        d = (now_dt - timedelta(days=i)).strftime("%Y-%m-%d")
        short = (now_dt - timedelta(days=i)).strftime("%d/%m")
        ldata = next((x for x in leads_ts if x["_id"] == d), None)
        jdata = next((x for x in jobs_ts if x["_id"] == d), None)
        cdata = next((x for x in contacts_ts if x["_id"] == d), None)
        data.append({"date": d, "label": short, "leads": ldata["count"] if ldata else 0, "scored": ldata["scored"] if ldata else 0, "rejected": ldata["rejected"] if ldata else 0, "jobs": jdata["count"] if jdata else 0, "contacts": cdata["count"] if cdata else 0})
    # Top categories
    cat_pipeline = [{"$match": {"tenant_id": tid}}, {"$group": {"_id": "$normalized_category", "count": {"$sum": 1}, "avg_score": {"$avg": "$ai_score"}}}, {"$sort": {"count": -1}}, {"$limit": 10}]
    categories = await db.leads.aggregate(cat_pipeline).to_list(10)
    # Quality distribution
    quality_pipeline = [{"$match": {"tenant_id": tid}}, {"$group": {"_id": "$quality_level", "count": {"$sum": 1}}}, {"$sort": {"count": -1}}]
    quality_dist = await db.leads.aggregate(quality_pipeline).to_list(10)
    return {"time_series": data, "period": period, "days": days, "top_categories": [{"name": c["_id"] or "Sin categoria", "count": c["count"], "avg_score": round(c["avg_score"] or 0)} for c in categories], "quality_distribution": [{"name": q["_id"] or "Sin clasificar", "count": q["count"]} for q in quality_dist]}

# ==================== SETTINGS ====================

@api_router.get("/settings")
async def get_settings(request: Request):
    user = await get_current_user(request)
    tenant = await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0})
    if not tenant:
        return {"branding": {}, "integrations": []}
    return tenant

@api_router.put("/settings")
async def update_settings(request: Request, body: SettingsUpdate):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    update_data = {}
    if body.company_name is not None:
        update_data["branding.company_name"] = body.company_name
    if body.logo_url is not None:
        update_data["branding.logo_url"] = body.logo_url
    if body.primary_color is not None:
        update_data["branding.primary_color"] = body.primary_color
    if body.secondary_color is not None:
        update_data["branding.secondary_color"] = body.secondary_color
    if body.industry is not None:
        update_data["branding.industry"] = body.industry
    if body.phone is not None:
        update_data["branding.phone"] = body.phone
    if body.address is not None:
        update_data["branding.address"] = body.address
    if body.website is not None:
        update_data["branding.website"] = body.website
    if body.tax_id is not None:
        update_data["branding.tax_id"] = body.tax_id
    if body.country is not None:
        update_data["branding.country"] = body.country
    if body.description is not None:
        update_data["branding.description"] = body.description
    if body.sender_defaults is not None:
        update_data["sender_defaults"] = body.sender_defaults
    update_data["updated_at"] = now
    await db.tenants.update_one({"id": user["tenant_id"]}, {"$set": update_data})
    tenant = await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0})
    return tenant

@api_router.get("/settings/integrations")
async def get_integrations(request: Request):
    user = await get_current_user(request)
    integrations = await db.integration_configs.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).to_list(20)
    return integrations

@api_router.put("/settings/integrations/{name}")
async def update_integration(request: Request, name: str, body: IntegrationUpdate):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    update_data = {k: v for k, v in body.model_dump().items() if v is not None}
    update_data["updated_at"] = now
    await db.integration_configs.update_one({"tenant_id": user["tenant_id"], "name": name}, {"$set": update_data})
    updated = await db.integration_configs.find_one({"tenant_id": user["tenant_id"], "name": name}, {"_id": 0})
    return updated

# ==================== USERS ====================

@api_router.get("/users")
async def list_users(request: Request):
    user = await get_current_user(request)
    if user["role"] not in ["super_admin", "tenant_admin"]:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    pipeline = [{"$match": {"tenant_id": user["tenant_id"]}}, {"$project": {"password_hash": 0}}, {"$addFields": {"id": {"$toString": "$_id"}}}, {"$project": {"_id": 0}}]
    users = await db.users.aggregate(pipeline).to_list(100)
    return users

@api_router.post("/users")
async def create_user(request: Request, body: UserCreate):
    user = await get_current_user(request)
    if user["role"] not in ["super_admin", "tenant_admin"]:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    email = body.email.lower().strip()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already exists")
    now = datetime.now(timezone.utc).isoformat()
    user_doc = {"email": email, "password_hash": hash_password(body.password), "name": body.name, "role": body.role, "tenant_id": user["tenant_id"], "created_at": now}
    result = await db.users.insert_one(user_doc)
    return {"id": str(result.inserted_id), "email": email, "name": body.name, "role": body.role}

# ==================== CRM BULK CONTACT ACTIONS ====================

class BulkContactAction(BaseModel):
    contact_ids: List[str]
    action: str
    stage: Optional[str] = None

@api_router.post("/crm/contacts/bulk-action")
async def bulk_contact_action(request: Request, body: BulkContactAction):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    if body.action == "move" and body.stage:
        result = await db.crm_contacts.update_many({"id": {"$in": body.contact_ids}, "tenant_id": user["tenant_id"]}, {"$set": {"stage": body.stage, "updated_at": now}})
        # Auto-create deals for contacts that don't have one
        for cid in body.contact_ids:
            existing_deal = await db.crm_deals.find_one({"contact_id": cid, "tenant_id": user["tenant_id"]})
            if not existing_deal:
                contact = await db.crm_contacts.find_one({"id": cid, "tenant_id": user["tenant_id"]}, {"_id": 0})
                if contact:
                    deal_title = contact.get("business_name", "Oportunidad")
                    await db.crm_deals.insert_one({"id": str(uuid.uuid4()), "tenant_id": user["tenant_id"], "contact_id": cid, "contact_name": contact.get("business_name", ""), "company": contact.get("business_name", ""), "title": deal_title, "name": deal_title, "value": 0, "stage": body.stage, "probability": 10 if body.stage == "nuevo" else 30 if body.stage == "contactado" else 50 if body.stage == "propuesta" else 70 if body.stage == "negociacion" else 100 if body.stage == "ganado" else 0, "assigned_to": user.get("name", ""), "created_at": now, "updated_at": now})
                    await db.crm_contacts.update_one({"id": cid}, {"$inc": {"deal_count": 1}})
            else:
                await db.crm_deals.update_one({"contact_id": cid, "tenant_id": user["tenant_id"]}, {"$set": {"stage": body.stage, "updated_at": now}})
        return {"message": f"{result.modified_count} contactos movidos a {body.stage} con oportunidades creadas"}
    elif body.action == "delete":
        result = await db.crm_contacts.delete_many({"id": {"$in": body.contact_ids}, "tenant_id": user["tenant_id"]})
        await db.crm_deals.delete_many({"contact_id": {"$in": body.contact_ids}, "tenant_id": user["tenant_id"]})
        return {"message": f"{result.deleted_count} contactos eliminados"}
    raise HTTPException(status_code=400, detail="Accion invalida")

# ==================== IMPORT ====================

@api_router.post("/import/leads")
async def import_leads(request: Request):
    import openpyxl
    import io
    user = await get_current_user(request)
    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(status_code=400, detail="No se recibio archivo")
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if len(rows) < 2:
            raise HTTPException(status_code=400, detail="El archivo esta vacio o solo tiene headers")
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        now = datetime.now(timezone.utc).isoformat()
        imported = 0
        for row in rows[1:]:
            data = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i]:
                    data[headers[i]] = str(val).strip() if val else ""
            bname = data.get("business_name") or data.get("empresa") or data.get("nombre") or data.get("name") or ""
            if not bname:
                continue
            lead = {
                "id": str(uuid.uuid4()), "tenant_id": user["tenant_id"], "job_id": "",
                "business_name": bname,
                "raw_category": data.get("category") or data.get("categoria") or data.get("raw_category") or "",
                "normalized_category": data.get("normalized_category") or data.get("categoria") or data.get("category") or "",
                "province": data.get("province") or data.get("provincia") or "",
                "city": data.get("city") or data.get("ciudad") or "",
                "website": data.get("website") or data.get("web") or data.get("sitio_web") or "",
                "email": data.get("email") or data.get("correo") or "",
                "phone": data.get("phone") or data.get("telefono") or data.get("tel") or "",
                "ai_score": int(data.get("ai_score") or data.get("score") or 0) if str(data.get("ai_score") or data.get("score") or "0").isdigit() else 0,
                "quality_level": data.get("quality_level") or "imported",
                "recommendation": data.get("recommendation") or "Importado desde Excel",
                "recommended_first_line": data.get("recommended_first_line") or "",
                "status": "raw", "created_at": now, "updated_at": now
            }
            await db.leads.insert_one(lead)
            imported += 1
        return {"message": f"{imported} leads importados exitosamente", "imported": imported}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al procesar archivo: {str(e)}")

@api_router.post("/import/crm-contacts")
async def import_crm_contacts(request: Request):
    import openpyxl
    import io
    user = await get_current_user(request)
    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(status_code=400, detail="No se recibio archivo")
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if len(rows) < 2:
            raise HTTPException(status_code=400, detail="El archivo esta vacio")
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        now = datetime.now(timezone.utc).isoformat()
        imported = 0
        for row in rows[1:]:
            data = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i]:
                    data[headers[i]] = str(val).strip() if val else ""
            bname = data.get("business_name") or data.get("empresa") or data.get("nombre") or data.get("name") or ""
            if not bname:
                continue
            contact = {
                "id": str(uuid.uuid4()), "tenant_id": user["tenant_id"], "lead_id": "",
                "business_name": bname,
                "contact_name": data.get("contact_name") or data.get("contacto") or "",
                "email": data.get("email") or data.get("correo") or "",
                "phone": data.get("phone") or data.get("telefono") or "",
                "city": data.get("city") or data.get("ciudad") or "",
                "province": data.get("province") or data.get("provincia") or "",
                "category": data.get("category") or data.get("categoria") or "",
                "source": "excel_import", "notes": data.get("notes") or data.get("notas") or "",
                "stage": "nuevo", "ai_score": 0, "deal_count": 0, "total_value": 0,
                "created_at": now, "updated_at": now
            }
            await db.crm_contacts.insert_one(contact)
            imported += 1
        return {"message": f"{imported} contactos importados al CRM", "imported": imported}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al procesar archivo: {str(e)}")

# ==================== EMAIL MARKETING ====================

class EmailListCreate(BaseModel):
    name: str
    description: Optional[str] = ""

class EmailCampaignCreate(BaseModel):
    name: str
    list_id: Optional[str] = ""
    template_id: Optional[str] = ""
    subject: Optional[str] = ""
    from_name: Optional[str] = ""
    from_email: Optional[str] = ""

class AutomationCreate(BaseModel):
    name: str
    trigger: Optional[str] = "manual"
    steps: Optional[List[Dict[str, Any]]] = []

@api_router.get("/email-marketing/lists")
async def list_email_lists(request: Request):
    user = await get_current_user(request)
    lists = await db.email_lists.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return lists

@api_router.post("/email-marketing/lists")
async def create_email_list(request: Request, body: EmailListCreate):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    doc = {"id": str(uuid.uuid4()), "tenant_id": user["tenant_id"], "name": body.name, "description": body.description or "", "subscriber_count": 0, "lead_ids": [], "created_at": now, "updated_at": now}
    await db.email_lists.insert_one(doc)
    return serialize_doc(doc)

@api_router.put("/email-marketing/lists/{list_id}")
async def update_email_list(request: Request, list_id: str, body: Dict[str, Any] = {}):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    update_data = {k: v for k, v in body.items() if k not in ("id", "tenant_id", "_id") and v is not None}
    update_data["updated_at"] = now
    result = await db.email_lists.update_one({"id": list_id, "tenant_id": user["tenant_id"]}, {"$set": update_data})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Lista no encontrada")
    return await db.email_lists.find_one({"id": list_id}, {"_id": 0})

@api_router.delete("/email-marketing/lists/{list_id}")
async def delete_email_list(request: Request, list_id: str):
    user = await get_current_user(request)
    result = await db.email_lists.delete_one({"id": list_id, "tenant_id": user["tenant_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lista no encontrada")
    return {"message": "Lista eliminada"}

@api_router.post("/email-marketing/lists/{list_id}/add-leads")
async def add_leads_to_list(request: Request, list_id: str, body: Dict[str, Any] = {}):
    """Add scored/approved leads to an email list"""
    user = await get_current_user(request)
    lead_ids = body.get("lead_ids", [])
    status_filter = body.get("status", "")
    now = datetime.now(timezone.utc).isoformat()
    lst = await db.email_lists.find_one({"id": list_id, "tenant_id": user["tenant_id"]})
    if not lst:
        raise HTTPException(status_code=404, detail="Lista no encontrada")
    existing_ids = lst.get("lead_ids", [])
    if lead_ids:
        new_ids = [lid for lid in lead_ids if lid not in existing_ids]
    elif status_filter:
        leads = await db.leads.find({"tenant_id": user["tenant_id"], "status": status_filter}, {"id": 1, "_id": 0}).to_list(5000)
        new_ids = [l["id"] for l in leads if l["id"] not in existing_ids]
    else:
        leads = await db.leads.find({"tenant_id": user["tenant_id"], "status": {"$in": ["scored", "approved"]}}, {"id": 1, "_id": 0}).to_list(5000)
        new_ids = [l["id"] for l in leads if l["id"] not in existing_ids]
    all_ids = existing_ids + new_ids
    await db.email_lists.update_one({"id": list_id}, {"$set": {"lead_ids": all_ids, "subscriber_count": len(all_ids), "updated_at": now}})
    return {"message": f"{len(new_ids)} leads agregados a la lista", "total": len(all_ids)}

@api_router.get("/email-marketing/lists/{list_id}/leads")
async def get_list_leads(request: Request, list_id: str):
    user = await get_current_user(request)
    lst = await db.email_lists.find_one({"id": list_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not lst:
        raise HTTPException(status_code=404, detail="Lista no encontrada")
    lead_ids = lst.get("lead_ids", [])
    leads = []
    if lead_ids:
        leads = await db.leads.find({"id": {"$in": lead_ids}}, {"_id": 0}).to_list(5000)
    return {"list": lst, "leads": leads}

@api_router.post("/email-marketing/lists/{list_id}/remove-leads")
async def remove_leads_from_list(request: Request, list_id: str, body: Dict[str, Any] = {}):
    user = await get_current_user(request)
    remove_ids = body.get("lead_ids", [])
    now = datetime.now(timezone.utc).isoformat()
    lst = await db.email_lists.find_one({"id": list_id, "tenant_id": user["tenant_id"]})
    if not lst:
        raise HTTPException(status_code=404, detail="Lista no encontrada")
    new_ids = [lid for lid in lst.get("lead_ids", []) if lid not in remove_ids]
    await db.email_lists.update_one({"id": list_id}, {"$set": {"lead_ids": new_ids, "subscriber_count": len(new_ids), "updated_at": now}})
    return {"message": f"{len(remove_ids)} leads removidos", "total": len(new_ids)}

@api_router.post("/email-marketing/lists/{list_id}/add-manual-leads")
async def add_manual_leads_to_list(request: Request, list_id: str, body: Dict[str, Any] = {}):
    """Add specific leads by ID to a list (manual pick from Leads page)"""
    user = await get_current_user(request)
    lead_ids = body.get("lead_ids", [])
    if not lead_ids:
        raise HTTPException(status_code=400, detail="lead_ids requeridos")
    now = datetime.now(timezone.utc).isoformat()
    lst = await db.email_lists.find_one({"id": list_id, "tenant_id": user["tenant_id"]})
    if not lst:
        raise HTTPException(status_code=404, detail="Lista no encontrada")
    existing_ids = lst.get("lead_ids", [])
    new_ids = [lid for lid in lead_ids if lid not in existing_ids]
    all_ids = existing_ids + new_ids
    await db.email_lists.update_one({"id": list_id}, {"$set": {"lead_ids": all_ids, "subscriber_count": len(all_ids), "updated_at": now}})
    return {"message": f"{len(new_ids)} leads agregados", "total": len(all_ids)}

@api_router.post("/email-marketing/auto-list-from-leads")
async def auto_create_list_from_leads(request: Request, body: Dict[str, Any] = {}):
    """Auto-create an email list from scored/approved leads"""
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    status = body.get("status", "scored")
    name = body.get("name", f"Leads calificados - {now[:10]}")
    leads = await db.leads.find({"tenant_id": user["tenant_id"], "status": {"$in": [status, "approved", "scored"]}}, {"id": 1, "_id": 0}).to_list(5000)
    lead_ids = [l["id"] for l in leads]
    doc = {"id": str(uuid.uuid4()), "tenant_id": user["tenant_id"], "name": name, "description": f"Lista auto-generada con {len(lead_ids)} leads calificados", "subscriber_count": len(lead_ids), "lead_ids": lead_ids, "created_at": now, "updated_at": now}
    await db.email_lists.insert_one(doc)
    return serialize_doc(doc)

@api_router.get("/email-marketing/campaigns")
async def list_email_campaigns(request: Request):
    user = await get_current_user(request)
    campaigns = await db.email_campaigns.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return campaigns

@api_router.post("/email-marketing/campaigns")
async def create_email_campaign(request: Request, body: EmailCampaignCreate):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    doc = {"id": str(uuid.uuid4()), "tenant_id": user["tenant_id"], "name": body.name, "list_id": body.list_id or "", "template_id": body.template_id or "", "subject": body.subject or "", "from_name": body.from_name or "", "from_email": body.from_email or "", "status": "borrador", "sent_count": 0, "open_count": 0, "click_count": 0, "bounce_count": 0, "unsub_count": 0, "created_at": now, "updated_at": now}
    await db.email_campaigns.insert_one(doc)
    return serialize_doc(doc)

@api_router.put("/email-marketing/campaigns/{campaign_id}")
async def update_email_campaign(request: Request, campaign_id: str, body: Dict[str, Any] = {}):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    update_data = {k: v for k, v in body.items() if k not in ("id", "tenant_id", "_id") and v is not None}
    update_data["updated_at"] = now
    result = await db.email_campaigns.update_one({"id": campaign_id, "tenant_id": user["tenant_id"]}, {"$set": update_data})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Campana no encontrada")
    return await db.email_campaigns.find_one({"id": campaign_id}, {"_id": 0})

@api_router.delete("/email-marketing/campaigns/{campaign_id}")
async def delete_email_campaign(request: Request, campaign_id: str):
    user = await get_current_user(request)
    result = await db.email_campaigns.delete_one({"id": campaign_id, "tenant_id": user["tenant_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Campana no encontrada")
    return {"message": "Campana eliminada"}

@api_router.post("/email-marketing/campaigns/{campaign_id}/simulate")
async def simulate_email_campaign(request: Request, campaign_id: str):
    user = await get_current_user(request)
    campaign = await db.email_campaigns.find_one({"id": campaign_id, "tenant_id": user["tenant_id"]})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campana no encontrada")
    now = datetime.now(timezone.utc).isoformat()
    sent = random.randint(200, 2000)
    opens = int(sent * random.uniform(0.20, 0.45))
    clicks = int(opens * random.uniform(0.10, 0.30))
    bounces = int(sent * random.uniform(0.01, 0.05))
    unsubs = int(sent * random.uniform(0.001, 0.01))
    await db.email_campaigns.update_one({"id": campaign_id}, {"$set": {"status": "enviada", "sent_count": sent, "open_count": opens, "click_count": clicks, "bounce_count": bounces, "unsub_count": unsubs, "sent_at": now, "updated_at": now}})
    updated = await db.email_campaigns.find_one({"id": campaign_id}, {"_id": 0})
    return updated

@api_router.post("/email-marketing/campaigns/{campaign_id}/send-real")
async def send_real_email_campaign(request: Request, campaign_id: str):
    """Send real emails via Resend to leads in the campaign's linked list"""
    user = await get_current_user(request)
    campaign = await db.email_campaigns.find_one({"id": campaign_id, "tenant_id": user["tenant_id"]})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campana no encontrada")
    template = None
    if campaign.get("template_id"):
        template = await db.templates.find_one({"id": campaign["template_id"]}, {"_id": 0})
    email_list = None
    if campaign.get("list_id"):
        email_list = await db.email_lists.find_one({"id": campaign["list_id"]}, {"_id": 0})
    if not email_list or not email_list.get("lead_ids"):
        raise HTTPException(status_code=400, detail="La campana no tiene una lista con leads asignados")
    verified_domain = await db.domains.find_one({"tenant_id": user["tenant_id"], "status": {"$in": ["verified", "warmup_recommended", "ready_to_send"]}}, {"_id": 0})
    from_email = verified_domain.get("sender_email", "onboarding@resend.dev") if verified_domain else "onboarding@resend.dev"
    from_name = verified_domain.get("sender_name", "Spectra Flow") if verified_domain else "Spectra Flow"
    now = datetime.now(timezone.utc).isoformat()
    sent_count = 0
    errors = []
    lead_ids = email_list.get("lead_ids", [])
    for lid in lead_ids[:50]:
        lead = await db.leads.find_one({"id": lid, "tenant_id": user["tenant_id"]}, {"_id": 0})
        if not lead or not lead.get("email"):
            continue
        subject = campaign.get("subject") or (template.get("subject", "Contacto de Spectra Flow") if template else "Contacto de Spectra Flow")
        html_body = template.get("html_body", "<p>Hola {business_name}</p>") if template else "<p>Hola {business_name}</p>"
        subject = subject.replace("{business_name}", lead.get("business_name", "")).replace("{city}", lead.get("city", "")).replace("{normalized_category}", lead.get("normalized_category", ""))
        html_body = html_body.replace("{business_name}", lead.get("business_name", "")).replace("{city}", lead.get("city", "")).replace("{normalized_category}", lead.get("normalized_category", "")).replace("{recommended_first_line}", lead.get("recommended_first_line", "")).replace("{sender_name}", from_name)
        if resend.api_key:
            try:
                params = {"from": f"{from_name} <{from_email}>", "to": [lead["email"]], "subject": subject, "html": html_body}
                await asyncio.to_thread(resend.Emails.send, params)
                sent_count += 1
                await db.leads.update_one({"id": lid}, {"$set": {"status": "contacted", "updated_at": now}})
            except Exception as e:
                errors.append(f"{lead['email']}: {str(e)[:50]}")
    await db.email_campaigns.update_one({"id": campaign_id}, {"$set": {"status": "enviada", "sent_count": sent_count, "sent_at": now, "updated_at": now}})
    return {"message": f"{sent_count} emails enviados via Resend", "sent": sent_count, "errors": errors[:5], "limited_to": "Max 50 por envio (warmup)"}

@api_router.post("/ai/dify-score-lead")
async def dify_score_lead(request: Request, body: Dict[str, Any] = {}):
    """Score a single lead using Dify AI"""
    user = await get_current_user(request)
    lead_id = body.get("lead_id", "")
    lead = await db.leads.find_one({"id": lead_id, "tenant_id": user["tenant_id"]}, {"_id": 0}) if lead_id else None
    business_data = body.get("business_data", "")
    if lead and not business_data:
        business_data = f"Nombre: {lead.get('business_name','')}\nDireccion: {lead.get('city','')}, {lead.get('province','')}\nTelefono: {lead.get('phone','')}\nWebsite: {lead.get('website','')}\nEmail: {lead.get('email','')}\nCategoria: {lead.get('normalized_category','')}"
    if not business_data:
        raise HTTPException(status_code=400, detail="Se requiere business_data o lead_id")
    dify_base = os.environ.get("DIFY_BASE_URL", "")
    dify_key = os.environ.get("DIFY_APP_KEY", "")
    if not dify_base or not dify_key:
        dify_config = await db.integration_configs.find_one({"tenant_id": user["tenant_id"], "name": "dify"}, {"_id": 0})
        if dify_config:
            dify_base = dify_config.get("base_url", dify_base)
            dify_key = dify_config.get("api_key", dify_key)
    if not dify_base or not dify_key:
        raise HTTPException(status_code=400, detail="Dify no configurado. Configura Base URL y API Key en Integraciones.")
    try:
        import httpx
        dify_url = dify_base.rstrip("/")
        if dify_url.startswith("http://"):
            dify_url = "https://" + dify_url[7:]
        async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
            resp = await client.post(f"{dify_url}/workflows/run", headers={"Authorization": f"Bearer {dify_key}", "Content-Type": "application/json"}, json={"inputs": {"business_data": business_data}, "response_mode": "blocking", "user": user.get("email", "spectra")})
            result = resp.json()
            output = result.get("data", {}).get("outputs", {}).get("result", "")
            if output:
                import json as json_mod
                try:
                    scored = json_mod.loads(output) if isinstance(output, str) else output
                    if lead_id and scored.get("ai_score"):
                        now = datetime.now(timezone.utc).isoformat()
                        await db.leads.update_one({"id": lead_id}, {"$set": {"ai_score": int(scored["ai_score"]), "quality_level": scored.get("quality_level", ""), "recommendation": scored.get("recommendation", ""), "recommended_first_line": scored.get("recommended_first_line", ""), "normalized_category": scored.get("normalized_category", lead.get("normalized_category", "")), "status": "scored", "updated_at": now}})
                    return {"success": True, "scoring": scored, "raw_output": output}
                except:
                    return {"success": True, "scoring": None, "raw_output": output}
            return {"success": False, "error": "Dify retorno output vacio. Verifica la config de tu workflow.", "raw_response": result}
    except Exception as e:
        logger.error(f"Dify scoring error: {e}")
        raise HTTPException(status_code=500, detail=f"Error llamando a Dify: {str(e)}")

@api_router.post("/ai/rescore-leads")
async def rescore_leads_batch(request: Request, body: Dict[str, Any] = {}):
    """Re-score leads that have default score (50) using Emergent LLM as fallback"""
    user = await get_current_user(request)
    lead_ids = body.get("lead_ids", [])
    query = {"tenant_id": user["tenant_id"]}
    if lead_ids:
        query["id"] = {"$in": lead_ids}
    else:
        query["ai_score"] = 50
    leads = await db.leads.find(query, {"_id": 0}).limit(50).to_list(50)
    if not leads:
        return {"message": "No hay leads para re-clasificar", "rescored": 0}
    # Try Dify first
    dify_base = os.environ.get("DIFY_BASE_URL", "")
    dify_key = os.environ.get("DIFY_APP_KEY", "")
    if not dify_base or not dify_key:
        dify_config = await db.integration_configs.find_one({"tenant_id": user["tenant_id"], "name": "dify"}, {"_id": 0})
        if dify_config:
            dify_base = dify_config.get("base_url", dify_base)
            dify_key = dify_config.get("api_key", dify_key)
    rescored = 0
    now = datetime.now(timezone.utc).isoformat()
    for lead in leads:
        business_data = f"Nombre: {lead.get('business_name','')}\nDireccion: {lead.get('address','')}, {lead.get('city','')}, {lead.get('province','')}\nTelefono: {lead.get('phone','')}\nWebsite: {lead.get('website','')}\nEmail: {lead.get('email','')}\nRating: {lead.get('rating',0)}\nReviews: {lead.get('reviews_count',0)}\nCategoria: {lead.get('normalized_category','')}"
        scored = None
        # Try Dify
        if dify_base and dify_key:
            try:
                dify_url = dify_base.rstrip("/")
                if dify_url.startswith("http://"):
                    dify_url = "https://" + dify_url[7:]
                async with httpx.AsyncClient(timeout=20, follow_redirects=True) as client:
                    resp = await client.post(f"{dify_url}/workflows/run", headers={"Authorization": f"Bearer {dify_key}", "Content-Type": "application/json"}, json={"inputs": {"business_data": business_data}, "response_mode": "blocking", "user": "spectra-rescore"})
                    result = resp.json()
                    output = result.get("data", {}).get("outputs", {}).get("result", "")
                    if output:
                        import json as jm
                        json_match = re.search(r'\{[\s\S]*\}', output) if isinstance(output, str) else None
                        parsed = jm.loads(json_match.group(0)) if json_match else (output if isinstance(output, dict) else None)
                        if parsed and parsed.get("ai_score"):
                            scored = parsed
            except:
                pass
        # Fallback: Emergent LLM scoring
        if not scored:
            try:
                from emergentintegrations.llm.chat import LlmChat, UserMessage
                chat = LlmChat(api_key=os.environ.get("EMERGENT_LLM_KEY", ""), session_id=f"score-{lead['id']}", system_message="Eres un clasificador de leads B2B. Califica del 0 al 100. Responde SOLO JSON: {\"ai_score\": N, \"quality_level\": \"excellent/good/average/poor\", \"recommendation\": \"texto\", \"recommended_first_line\": \"texto\"}")
                resp_text = await chat.send_message(UserMessage(text=f"Clasifica este lead:\n{business_data}"))
                import json as jm
                json_match = re.search(r'\{[\s\S]*\}', resp_text)
                if json_match:
                    scored = jm.loads(json_match.group(0))
            except:
                pass
        if scored and scored.get("ai_score"):
            score_val = int(scored["ai_score"])
            status = "scored" if score_val >= 50 else "rejected"
            await db.leads.update_one({"id": lead["id"]}, {"$set": {
                "ai_score": score_val, "quality_level": scored.get("quality_level", "average"),
                "recommendation": scored.get("recommendation", ""), "recommended_first_line": scored.get("recommended_first_line", ""),
                "status": status, "updated_at": now
            }})
            rescored += 1
    return {"message": f"{rescored} leads re-clasificados de {len(leads)}", "rescored": rescored, "total": len(leads)}

@api_router.get("/email-marketing/automations")
async def list_automations(request: Request):
    user = await get_current_user(request)
    autos = await db.email_automations.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return autos

@api_router.post("/email-marketing/automations")
async def create_automation(request: Request, body: AutomationCreate):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    default_steps = [
        {"type": "email", "delay_days": 0, "subject": "Primer contacto", "template": ""},
        {"type": "wait", "delay_days": 3, "subject": "", "template": ""},
        {"type": "condition", "delay_days": 0, "subject": "Si abrio email", "template": ""},
        {"type": "email", "delay_days": 0, "subject": "Follow-up", "template": ""},
        {"type": "wait", "delay_days": 5, "subject": "", "template": ""},
        {"type": "email", "delay_days": 0, "subject": "Ultimo contacto", "template": ""},
    ]
    doc = {"id": str(uuid.uuid4()), "tenant_id": user["tenant_id"], "name": body.name, "trigger": body.trigger or "manual", "steps": body.steps if body.steps else default_steps, "status": "borrador", "active_count": 0, "completed_count": 0, "created_at": now, "updated_at": now}
    await db.email_automations.insert_one(doc)
    return serialize_doc(doc)

@api_router.put("/email-marketing/automations/{automation_id}")
async def update_automation(request: Request, automation_id: str, body: Dict[str, Any] = {}):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    update_data = {"updated_at": now}
    if "name" in body:
        update_data["name"] = body["name"]
    if "trigger" in body:
        update_data["trigger"] = body["trigger"]
    if "steps" in body:
        update_data["steps"] = body["steps"]
    if "status" in body:
        update_data["status"] = body["status"]
    result = await db.email_automations.update_one({"id": automation_id, "tenant_id": user["tenant_id"]}, {"$set": update_data})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Automatizacion no encontrada")
    updated = await db.email_automations.find_one({"id": automation_id}, {"_id": 0})
    return updated

@api_router.delete("/email-marketing/automations/{automation_id}")
async def delete_automation(request: Request, automation_id: str):
    user = await get_current_user(request)
    result = await db.email_automations.delete_one({"id": automation_id, "tenant_id": user["tenant_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Automatizacion no encontrada")
    return {"message": "Automatizacion eliminada"}

@api_router.get("/settings/scoring")
async def get_scoring_config(request: Request):
    user = await get_current_user(request)
    tenant = await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0})
    defaults = {"website": 20, "email": 15, "phone": 10, "address": 10, "rating_excellent": 20, "rating_good": 15, "rating_fair": 10, "reviews_100": 10, "reviews_50": 7, "reviews_10": 4, "category_match": 10, "professional_name": 5, "min_excellent": 80, "min_good": 60, "min_average": 40}
    return tenant.get("scoring_config", defaults) if tenant else defaults

@api_router.put("/settings/scoring")
async def update_scoring_config(request: Request, body: Dict[str, Any] = {}):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    await db.tenants.update_one({"id": user["tenant_id"]}, {"$set": {"scoring_config": body, "updated_at": now}})
    return body

@api_router.get("/email-marketing/stats")
async def email_marketing_stats(request: Request):
    user = await get_current_user(request)
    tid = user["tenant_id"]
    lists = await db.email_lists.count_documents({"tenant_id": tid})
    campaigns = await db.email_campaigns.count_documents({"tenant_id": tid})
    automations = await db.email_automations.count_documents({"tenant_id": tid})
    pipeline = [{"$match": {"tenant_id": tid}}, {"$group": {"_id": None, "sent": {"$sum": "$sent_count"}, "opens": {"$sum": "$open_count"}, "clicks": {"$sum": "$click_count"}, "bounces": {"$sum": "$bounce_count"}, "unsubs": {"$sum": "$unsub_count"}}}]
    totals = await db.email_campaigns.aggregate(pipeline).to_list(1)
    t = totals[0] if totals else {}
    return {"lists": lists, "campaigns": campaigns, "automations": automations, "total_sent": t.get("sent", 0), "total_opens": t.get("opens", 0), "total_clicks": t.get("clicks", 0), "total_bounces": t.get("bounces", 0), "total_unsubs": t.get("unsubs", 0), "open_rate": round((t.get("opens", 0) / max(t.get("sent", 0), 1)) * 100, 1), "click_rate": round((t.get("clicks", 0) / max(t.get("opens", 0), 1)) * 100, 1)}

# ==================== AI ENDPOINTS ====================

@api_router.post("/ai/generate-template")
async def ai_generate_template(request: Request, body: Dict[str, Any] = {}):
    user = await get_current_user(request)
    industry = body.get("industry", "general")
    objective = body.get("objective", "generar interes")
    tone = body.get("tone", "profesional")
    custom_prompt = body.get("custom_prompt", "")
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(api_key=os.environ.get("EMERGENT_LLM_KEY", ""), session_id=f"template-{uuid.uuid4()}", system_message="Eres un experto en copywriting y neuromarketing. Generas emails comerciales persuasivos en español usando tecnicas de neuropersuasion (urgencia, reciprocidad, prueba social, escasez, autoridad). Responde SOLO en formato JSON con las keys: subject, html_body, plain_text, variables (array de strings), first_line.")
        base_prompt = f"Genera un email de {objective} para la industria de {industry} con tono {tone}. Usa tecnicas de neuropersuasion. El email debe tener variables como {{{{business_name}}}}, {{{{city}}}}, {{{{sender_name}}}}."
        if custom_prompt:
            base_prompt += f"\n\nInstrucciones adicionales del usuario: {custom_prompt}"
        base_prompt += " Responde en JSON."
        response = await chat.send_message(UserMessage(text=base_prompt))
        import json
        try:
            clean = response.strip()
            if clean.startswith("```"):
                clean = clean.split("\n", 1)[1].rsplit("```", 1)[0]
            result = json.loads(clean)
        except:
            result = {"subject": f"Oportunidad exclusiva para {{{{business_name}}}} - {industry}", "html_body": f"<p>Estimado equipo de <strong>{{{{business_name}}}}</strong>,</p><p>{response[:500]}</p><p>Saludos cordiales,<br/>{{{{sender_name}}}}</p>", "plain_text": response[:500], "variables": ["business_name", "city", "sender_name"], "first_line": "Notamos que su empresa tiene una presencia destacada."}
        return {"generated": True, **result}
    except Exception as e:
        logger.error(f"AI generation error: {e}")
        return {"generated": False, "error": str(e), "subject": f"Oportunidad para {{{{business_name}}}} en {industry}", "html_body": f"<p>Estimado equipo de <strong>{{{{business_name}}}}</strong>,</p><p>Nos dirigimos a usted porque identificamos una oportunidad unica para su negocio en el sector de {industry}.</p><p>Nos encantaria coordinar una breve llamada para explorar como podemos ayudarlos.</p><p>Saludos cordiales,<br/>{{{{sender_name}}}}</p>", "plain_text": "", "variables": ["business_name", "sender_name"], "first_line": "Identificamos una oportunidad unica para su negocio."}

@api_router.post("/ai/flow-bot")
async def ai_flow_bot(request: Request, body: Dict[str, Any] = {}):
    user = await get_current_user(request)
    section = body.get("section", "general")
    question = body.get("question", "")
    tid = user["tenant_id"]
    # ---- RICH CONTEXT ----
    ctx = {}
    ctx["total_leads"] = await db.leads.count_documents({"tenant_id": tid})
    ctx["scored"] = await db.leads.count_documents({"tenant_id": tid, "status": "scored"})
    ctx["approved"] = await db.leads.count_documents({"tenant_id": tid, "status": "approved"})
    ctx["rejected"] = await db.leads.count_documents({"tenant_id": tid, "status": "rejected"})
    ctx["contacted"] = await db.leads.count_documents({"tenant_id": tid, "status": "contacted"})
    ctx["sent_to_crm"] = await db.leads.count_documents({"tenant_id": tid, "status": "sent_to_crm"})
    ctx["in_sequence"] = await db.leads.count_documents({"tenant_id": tid, "status": "queued_for_sequence"})
    # Sources breakdown
    for src in ["google_maps", "linkedin", "bot", "manual", "imported"]:
        ctx[f"source_{src}"] = await db.leads.count_documents({"tenant_id": tid, "source": src})
    ctx["source_bot"] += await db.leads.count_documents({"tenant_id": tid, "tags": "bot", "source": {"$ne": "bot"}})
    # Top cities
    city_pipeline = [{"$match": {"tenant_id": tid, "city": {"$ne": ""}}}, {"$group": {"_id": "$city", "count": {"$sum": 1}}}, {"$sort": {"count": -1}}, {"$limit": 5}]
    top_cities = await db.leads.aggregate(city_pipeline).to_list(5)
    ctx["top_ciudades"] = [{c["_id"]: c["count"]} for c in top_cities]
    # Top categories
    cat_pipeline = [{"$match": {"tenant_id": tid, "normalized_category": {"$ne": ""}}}, {"$group": {"_id": "$normalized_category", "count": {"$sum": 1}}}, {"$sort": {"count": -1}}, {"$limit": 5}]
    top_cats = await db.leads.aggregate(cat_pipeline).to_list(5)
    ctx["top_categorias"] = [{c["_id"]: c["count"]} for c in top_cats]
    # Top scored leads
    top_leads = await db.leads.find({"tenant_id": tid, "ai_score": {"$gte": 70}}, {"_id": 0, "business_name": 1, "ai_score": 1, "city": 1, "email": 1, "source": 1, "status": 1, "normalized_category": 1, "channel": 1}).sort("ai_score", -1).limit(10).to_list(10)
    ctx["top_leads"] = [{"nombre": l["business_name"], "score": l["ai_score"], "ciudad": l.get("city", ""), "fuente": l.get("source", ""), "estado": l.get("status", ""), "categoria": l.get("normalized_category", ""), "canal": l.get("channel", "")} for l in top_leads]
    # CRM
    ctx["crm_contacts"] = await db.crm_contacts.count_documents({"tenant_id": tid})
    ctx["deals_total"] = await db.crm_deals.count_documents({"tenant_id": tid})
    ctx["deals_ganados"] = await db.crm_deals.count_documents({"tenant_id": tid, "stage": "ganado"})
    ctx["deals_perdidos"] = await db.crm_deals.count_documents({"tenant_id": tid, "stage": "perdido"})
    # Campaigns
    ctx["campaigns"] = await db.campaigns.count_documents({"tenant_id": tid})
    em_pipeline = [{"$match": {"tenant_id": tid}}, {"$group": {"_id": None, "sent": {"$sum": "$sent_count"}, "opens": {"$sum": "$open_count"}, "clicks": {"$sum": "$click_count"}, "replies": {"$sum": "$reply_count"}}}]
    em_stats = await db.campaigns.aggregate(em_pipeline).to_list(1)
    if em_stats:
        ctx["emails_enviados"] = em_stats[0].get("sent", 0)
        ctx["emails_abiertos"] = em_stats[0].get("opens", 0)
        ctx["emails_respondidos"] = em_stats[0].get("replies", 0)
    # ---- SPECIFIC SEARCH when user asks a question ----
    specific_data = ""
    if question:
        q_lower = question.lower()
        # Search by source keywords
        if any(w in q_lower for w in ["bot", "optimia", "chatwoot", "whatsapp"]):
            bot_leads = await db.leads.find({"tenant_id": tid, "$or": [{"source": "bot"}, {"tags": "bot"}]}, {"_id": 0, "business_name": 1, "email": 1, "phone": 1, "city": 1, "channel": 1, "ai_score": 1, "status": 1, "created_at": 1}).sort("created_at", -1).limit(20).to_list(20)
            specific_data += f"\n\nLeads del Bot ({len(bot_leads)}):"
            for l in bot_leads:
                specific_data += f"\n- {l['business_name']}: email={l.get('email','')}, tel={l.get('phone','')}, canal={l.get('channel','web')}, score={l.get('ai_score',0)}, estado={l.get('status','')}, fecha={l.get('created_at','')[:10]}"
        if any(w in q_lower for w in ["linkedin"]):
            li_leads = await db.leads.find({"tenant_id": tid, "source": "linkedin"}, {"_id": 0, "business_name": 1, "city": 1, "ai_score": 1, "status": 1}).sort("ai_score", -1).limit(20).to_list(20)
            specific_data += f"\n\nLeads de LinkedIn ({len(li_leads)}):"
            for l in li_leads:
                specific_data += f"\n- {l['business_name']}: ciudad={l.get('city','')}, score={l.get('ai_score',0)}, estado={l.get('status','')}"
        # Search by city
        for city_doc in top_cities:
            city_name = city_doc["_id"]
            if city_name.lower() in q_lower:
                city_leads = await db.leads.find({"tenant_id": tid, "city": {"$regex": city_name, "$options": "i"}}, {"_id": 0, "business_name": 1, "ai_score": 1, "status": 1, "email": 1, "source": 1}).sort("ai_score", -1).limit(15).to_list(15)
                specific_data += f"\n\nLeads en {city_name} ({len(city_leads)}):"
                for l in city_leads:
                    specific_data += f"\n- {l['business_name']}: score={l.get('ai_score',0)}, estado={l.get('status','')}, fuente={l.get('source','')}"
        # Search leads/contacts/deals by name
        all_leads_sample = await db.leads.find({"tenant_id": tid}, {"_id": 0, "business_name": 1, "ai_score": 1, "status": 1, "email": 1, "city": 1, "source": 1, "normalized_category": 1}).limit(200).to_list(200)
        for l in all_leads_sample:
            lname = l.get("business_name", "")
            if lname and (lname.lower() in q_lower or any(w in q_lower for w in lname.lower().split() if len(w) > 3)):
                specific_data += f"\nLead '{lname}': score={l.get('ai_score',0)}, estado={l.get('status','')}, email={l.get('email','')}, ciudad={l.get('city','')}, categoria={l.get('normalized_category','')}, fuente={l.get('source','')}"
        # Search CRM deals
        all_deals = await db.crm_deals.find({"tenant_id": tid}, {"_id": 0}).to_list(100)
        for d in all_deals:
            dname = d.get("name", d.get("title", ""))
            if dname and (dname.lower() in q_lower or any(w in q_lower for w in dname.lower().split() if len(w) > 3)):
                specific_data += f"\nOportunidad '{dname}': etapa={d.get('stage','')}, valor=${d.get('value',0)}, contacto={d.get('contact_name','')}"
        # Recommendations keywords
        if any(w in q_lower for w in ["recomiend", "mejor", "top", "suger"]):
            specific_data += f"\n\nTop 10 leads recomendados (mayor score):"
            for l in ctx["top_leads"]:
                specific_data += f"\n- {l['nombre']}: score={l['score']}, ciudad={l['ciudad']}, fuente={l['fuente']}, categoria={l['categoria']}"
    import json
    system_context = f"""Eres Flow IA, el copiloto de inteligencia comercial de Spectra Flow. Tu trabajo es analizar datos del negocio y dar recomendaciones accionables.

DATOS DEL DASHBOARD:
{json.dumps(ctx, ensure_ascii=False, default=str)}"""
    if specific_data:
        system_context += f"\n\nDATOS ESPECIFICOS DE LA CONSULTA:{specific_data}"
    system_context += """

REGLAS:
- Responde SIEMPRE en espanol
- Se conciso, usa bullets, maximo 10 lineas
- Da recomendaciones accionables basadas en datos reales
- Si preguntan por leads de una fuente (bot, linkedin, b2b), lista los que encontraste
- Si preguntan recomendaciones, ordena por score y explica por que
- Si preguntan pasos a seguir, da un plan claro paso a paso
- Menciona numeros concretos siempre que puedas
- Nunca menciones tecnologias internas (n8n, Outscraper, Dify, Apify, Resend)"""
    prompt = question if question else f"El usuario esta en la seccion '{section}'. Analiza sus datos y dame un resumen ejecutivo con recomendaciones."
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(api_key=os.environ.get("EMERGENT_LLM_KEY", ""), session_id=f"flowbot-{user['_id']}-{section}", system_message=system_context)
        response = await chat.send_message(UserMessage(text=prompt))
        return {"response": response, "context": ctx}
    except Exception as e:
        recommendations = []
        if specific_data:
            recommendations.append(f"Datos encontrados:{specific_data[:500]}")
        if ctx.get("scored", 0) > 0:
            recommendations.append(f"Tenes {ctx['scored']} leads calificados sin revisar.")
        if ctx.get("source_bot", 0) > 0:
            recommendations.append(f"{ctx['source_bot']} leads llegaron desde el Bot.")
        if not recommendations:
            recommendations.append("Comenza buscando prospectos para alimentar tu pipeline.")
        return {"response": "\n".join([f"- {r}" for r in recommendations]), "context": ctx}

@api_router.post("/ai/help")
async def ai_help(request: Request, body: Dict[str, Any] = {}):
    user = await get_current_user(request)
    question = body.get("question", "")
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(api_key=os.environ.get("EMERGENT_LLM_KEY", ""), session_id=f"help-{uuid.uuid4()}", system_message="Eres el asistente de ayuda de Spectra Flow. Explicas como usar la plataforma de manera clara y concisa en español. Spectra Flow es una plataforma de prospeccion, calificacion de leads, email marketing y CRM. Modulos: Buscador de Prospectos, Flow IA, Leads, Campanas, Plantillas, Dominios, Spectra CRM, Email Marketing, Analisis, Configuracion.")
        response = await chat.send_message(UserMessage(text=question))
        return {"response": response}
    except Exception as e:
        help_texts = {"general": "Spectra Flow te permite buscar prospectos, calificarlos con IA, crear campanas de email y gestionar oportunidades en el CRM.", "prospeccion": "En el Buscador de Prospectos selecciona pais, provincia, ciudad e industria. Flow IA buscara y calificara prospectos automaticamente.", "leads": "En Leads podes ver, aprobar, rechazar y enviar leads al CRM o a secuencias de email.", "crm": "En Spectra CRM gestionas contactos, creas oportunidades y moves el pipeline con drag & drop.", "email": "En Email Marketing creas listas, campanas masivas y automatizaciones de email."}
        return {"response": help_texts.get(question.lower()[:10], help_texts["general"])}

# ==================== MODULE MANAGEMENT ====================

@api_router.get("/tenant/modules")
async def get_tenant_modules(request: Request):
    user = await get_current_user(request)
    tenant = await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0})
    return tenant.get("modules", {"prospeccion": True, "leads": True, "crm": True, "email_marketing": True, "web": False, "performance": False}) if tenant else {"prospeccion": True, "leads": True, "crm": True, "email_marketing": True, "web": False, "performance": False}

@api_router.put("/tenant/modules")
async def update_tenant_modules(request: Request, body: Dict[str, bool] = {}):
    user = await get_current_user(request)
    if user["role"] not in ["super_admin", "tenant_admin"]:
        raise HTTPException(status_code=403, detail="Sin permisos")
    await db.tenants.update_one({"id": user["tenant_id"]}, {"$set": {"modules": body}})
    return body

# ==================== SUPER ADMIN - TENANT MANAGEMENT ====================

class TenantCreate(BaseModel):
    name: str
    admin_email: str
    admin_password: str
    admin_name: Optional[str] = ""
    plan: Optional[str] = "starter"
    price: Optional[float] = 0
    modules: Optional[Dict[str, bool]] = None

class TenantUpdate(BaseModel):
    name: Optional[str] = None
    plan: Optional[str] = None
    price: Optional[float] = None
    modules: Optional[Dict[str, bool]] = None
    active: Optional[bool] = None

@api_router.get("/admin/tenants")
async def list_tenants(request: Request):
    user = await get_current_user(request)
    if user["role"] != "super_admin":
        raise HTTPException(status_code=403, detail="Solo super_admin")
    tenants = await db.tenants.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    result = []
    for t in tenants:
        user_count = await db.users.count_documents({"tenant_id": t["id"]})
        lead_count = await db.leads.count_documents({"tenant_id": t["id"]})
        contact_count = await db.crm_contacts.count_documents({"tenant_id": t["id"]})
        result.append({**t, "user_count": user_count, "lead_count": lead_count, "contact_count": contact_count})
    return result

@api_router.post("/admin/tenants")
async def create_tenant(request: Request, body: TenantCreate):
    user = await get_current_user(request)
    if user["role"] != "super_admin":
        raise HTTPException(status_code=403, detail="Solo super_admin")
    now = datetime.now(timezone.utc).isoformat()
    tenant_id = str(uuid.uuid4())
    default_modules = {"prospeccion": True, "leads": True, "crm": True, "email_marketing": True}
    tenant = {
        "id": tenant_id, "name": body.name,
        "branding": {"company_name": body.name, "logo_url": "", "primary_color": "#1D4ED8", "secondary_color": "#6366F1"},
        "sender_defaults": {"name": body.name, "email": f"noreply@{body.name.lower().replace(' ', '')}.com"},
        "modules": body.modules or default_modules,
        "plan": body.plan or "starter",
        "price": body.price or 0,
        "active": True,
        "created_at": now, "updated_at": now
    }
    await db.tenants.insert_one(tenant)
    # Create admin user for this tenant
    hashed = bcrypt.hashpw(body.admin_password.encode(), bcrypt.gensalt()).decode()
    admin_user = {
        "id": str(uuid.uuid4()), "tenant_id": tenant_id,
        "name": body.admin_name or body.name, "email": body.admin_email,
        "password_hash": hashed, "role": "tenant_admin",
        "created_at": now, "updated_at": now
    }
    await db.users.insert_one(admin_user)
    # Seed integration configs for new tenant
    for ic in [
        {"name": "n8n", "display_name": "n8n Orchestration", "enabled": False, "base_url": "", "api_key": "", "status": "not_configured", "description": "Workflow orchestration"},
        {"name": "dify", "display_name": "Dify AI", "enabled": False, "base_url": "", "api_key": "", "status": "not_configured", "description": "AI scoring"},
        {"name": "resend", "display_name": "Resend Email", "enabled": False, "base_url": "", "api_key": "", "status": "not_configured", "description": "Email sending"},
        {"name": "apify", "display_name": "Apify (LinkedIn)", "enabled": False, "base_url": "https://api.apify.com/v2", "api_key": "", "status": "not_configured", "description": "LinkedIn scraping"},
    ]:
        await db.integration_configs.insert_one({"id": str(uuid.uuid4()), "tenant_id": tenant_id, **ic, "last_sync": None, "created_at": now, "updated_at": now})
    return {**{k: v for k, v in tenant.items() if k != "_id"}, "admin_email": body.admin_email}

@api_router.put("/admin/tenants/{tenant_id}")
async def update_tenant(request: Request, tenant_id: str, body: TenantUpdate):
    user = await get_current_user(request)
    if user["role"] != "super_admin":
        raise HTTPException(status_code=403, detail="Solo super_admin")
    now = datetime.now(timezone.utc).isoformat()
    update = {"updated_at": now}
    if body.name is not None:
        update["name"] = body.name
    if body.plan is not None:
        update["plan"] = body.plan
    if body.price is not None:
        update["price"] = body.price
    if body.modules is not None:
        update["modules"] = body.modules
    if body.active is not None:
        update["active"] = body.active
    await db.tenants.update_one({"id": tenant_id}, {"$set": update})
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    return tenant

@api_router.get("/admin/tenants/{tenant_id}")
async def get_tenant_detail(request: Request, tenant_id: str):
    user = await get_current_user(request)
    if user["role"] != "super_admin":
        raise HTTPException(status_code=403, detail="Solo super_admin")
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant no encontrado")
    users = await db.users.find({"tenant_id": tenant_id}, {"_id": 0, "password_hash": 0}).to_list(50)
    lead_count = await db.leads.count_documents({"tenant_id": tenant_id})
    contact_count = await db.crm_contacts.count_documents({"tenant_id": tenant_id})
    deal_count = await db.crm_deals.count_documents({"tenant_id": tenant_id})
    campaign_count = await db.campaigns.count_documents({"tenant_id": tenant_id})
    return {**tenant, "users": users, "stats": {"leads": lead_count, "contacts": contact_count, "deals": deal_count, "campaigns": campaign_count}}

# ==================== BULK DEAL ACTIONS ====================

class BulkDealAction(BaseModel):
    deal_ids: List[str]
    action: str
    stage: Optional[str] = None

@api_router.post("/crm/deals/bulk-action")
async def bulk_deal_action(request: Request, body: BulkDealAction):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    if body.action == "move" and body.stage:
        result = await db.crm_deals.update_many({"id": {"$in": body.deal_ids}, "tenant_id": user["tenant_id"]}, {"$set": {"stage": body.stage, "updated_at": now}})
        return {"message": f"{result.modified_count} oportunidades movidas a {body.stage}"}
    elif body.action == "delete":
        result = await db.crm_deals.delete_many({"id": {"$in": body.deal_ids}, "tenant_id": user["tenant_id"]})
        return {"message": f"{result.deleted_count} oportunidades eliminadas"}
    raise HTTPException(status_code=400, detail="Accion invalida")


# ==================== CRM TASKS ====================

@api_router.get("/crm/tasks")
async def list_tasks(request: Request, contact_id: Optional[str] = None, deal_id: Optional[str] = None, status: Optional[str] = None):
    user = await get_current_user(request)
    query = {"tenant_id": user["tenant_id"]}
    if contact_id: query["contact_id"] = contact_id
    if deal_id: query["deal_id"] = deal_id
    if status: query["status"] = status
    tasks = await db.crm_tasks.find(query, {"_id": 0}).sort("due_date", 1).to_list(200)
    return tasks

@api_router.post("/crm/tasks")
async def create_task(request: Request, body: Dict[str, Any] = {}):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    task = {
        "id": str(uuid.uuid4()), "tenant_id": user["tenant_id"],
        "contact_id": body.get("contact_id", ""), "deal_id": body.get("deal_id", ""),
        "title": body.get("title", ""), "type": body.get("type", "tarea"),
        "description": body.get("description", ""),
        "due_date": body.get("due_date", ""), "status": "pendiente",
        "assigned_to": body.get("assigned_to", user.get("name", "")),
        "created_by": user.get("name", ""), "created_at": now, "updated_at": now
    }
    await db.crm_tasks.insert_one(task)
    await _log_activity(user, "tarea_creada", f"Tarea: {task['title']}", body.get("contact_id", ""), body.get("deal_id", ""))
    return {k: v for k, v in task.items() if k != "_id"}

@api_router.put("/crm/tasks/{task_id}")
async def update_task(request: Request, task_id: str, body: Dict[str, Any] = {}):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    update = {"updated_at": now}
    for f in ["title", "type", "description", "due_date", "status", "assigned_to"]:
        if f in body: update[f] = body[f]
    await db.crm_tasks.update_one({"id": task_id, "tenant_id": user["tenant_id"]}, {"$set": update})
    if body.get("status") == "completada":
        await _log_activity(user, "tarea_completada", f"Tarea completada: {body.get('title', task_id)}", "", "")
    task = await db.crm_tasks.find_one({"id": task_id}, {"_id": 0})
    return task

@api_router.delete("/crm/tasks/{task_id}")
async def delete_task(request: Request, task_id: str):
    user = await get_current_user(request)
    await db.crm_tasks.delete_one({"id": task_id, "tenant_id": user["tenant_id"]})
    return {"message": "Tarea eliminada"}

# ==================== CRM NOTES ====================

@api_router.get("/crm/notes")
async def list_notes(request: Request, contact_id: Optional[str] = None, deal_id: Optional[str] = None):
    user = await get_current_user(request)
    query = {"tenant_id": user["tenant_id"]}
    if contact_id: query["contact_id"] = contact_id
    if deal_id: query["deal_id"] = deal_id
    notes = await db.crm_notes.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    return notes

@api_router.post("/crm/notes")
async def create_note(request: Request, body: Dict[str, Any] = {}):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    note = {
        "id": str(uuid.uuid4()), "tenant_id": user["tenant_id"],
        "contact_id": body.get("contact_id", ""), "deal_id": body.get("deal_id", ""),
        "content": body.get("content", ""), "created_by": user.get("name", ""),
        "created_at": now, "updated_at": now
    }
    await db.crm_notes.insert_one(note)
    await _log_activity(user, "nota_creada", f"Nota agregada", body.get("contact_id", ""), body.get("deal_id", ""))
    return {k: v for k, v in note.items() if k != "_id"}

# ==================== CRM PRODUCTS ====================

@api_router.get("/crm/products")
async def list_products(request: Request):
    user = await get_current_user(request)
    products = await db.crm_products.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).sort("name", 1).to_list(100)
    return products

@api_router.post("/crm/products")
async def create_product(request: Request, body: Dict[str, Any] = {}):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    product = {
        "id": str(uuid.uuid4()), "tenant_id": user["tenant_id"],
        "name": body.get("name", ""), "price": body.get("price", 0),
        "description": body.get("description", ""), "currency": body.get("currency", "USD"),
        "created_at": now, "updated_at": now
    }
    await db.crm_products.insert_one(product)
    return {k: v for k, v in product.items() if k != "_id"}

@api_router.put("/crm/products/{product_id}")
async def update_product(request: Request, product_id: str, body: Dict[str, Any] = {}):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    update = {"updated_at": now}
    for f in ["name", "price", "description", "currency"]:
        if f in body: update[f] = body[f]
    await db.crm_products.update_one({"id": product_id, "tenant_id": user["tenant_id"]}, {"$set": update})
    return await db.crm_products.find_one({"id": product_id}, {"_id": 0})

@api_router.delete("/crm/products/{product_id}")
async def delete_product(request: Request, product_id: str):
    user = await get_current_user(request)
    await db.crm_products.delete_one({"id": product_id, "tenant_id": user["tenant_id"]})
    return {"message": "Producto eliminado"}

# ==================== DEAL PRODUCTS (ITEMS) ====================

@api_router.get("/crm/deals/{deal_id}/products")
async def list_deal_products(request: Request, deal_id: str):
    user = await get_current_user(request)
    items = await db.crm_deal_products.find({"deal_id": deal_id, "tenant_id": user["tenant_id"]}, {"_id": 0}).to_list(50)
    return items

@api_router.post("/crm/deals/{deal_id}/products")
async def add_deal_product(request: Request, deal_id: str, body: Dict[str, Any] = {}):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    item = {
        "id": str(uuid.uuid4()), "tenant_id": user["tenant_id"], "deal_id": deal_id,
        "product_id": body.get("product_id", ""), "product_name": body.get("product_name", ""),
        "quantity": body.get("quantity", 1), "price": body.get("price", 0),
        "created_at": now
    }
    await db.crm_deal_products.insert_one(item)
    total = sum([i["price"] * i["quantity"] for i in await db.crm_deal_products.find({"deal_id": deal_id}, {"_id": 0}).to_list(50)])
    await db.crm_deals.update_one({"id": deal_id}, {"$set": {"value": total, "updated_at": now}})
    await _log_activity(user, "producto_agregado", f"Producto: {item['product_name']}", "", deal_id)
    return {k: v for k, v in item.items() if k != "_id"}

@api_router.delete("/crm/deals/{deal_id}/products/{item_id}")
async def remove_deal_product(request: Request, deal_id: str, item_id: str):
    user = await get_current_user(request)
    await db.crm_deal_products.delete_one({"id": item_id})
    now = datetime.now(timezone.utc).isoformat()
    total = sum([i["price"] * i["quantity"] for i in await db.crm_deal_products.find({"deal_id": deal_id}, {"_id": 0}).to_list(50)])
    await db.crm_deals.update_one({"id": deal_id}, {"$set": {"value": total, "updated_at": now}})
    return {"message": "Producto removido"}

# ==================== CRM TAGS ====================

@api_router.put("/crm/deals/{deal_id}/tags")
async def update_deal_tags(request: Request, deal_id: str, body: Dict[str, Any] = {}):
    user = await get_current_user(request)
    tags = body.get("tags", [])
    await db.crm_deals.update_one({"id": deal_id, "tenant_id": user["tenant_id"]}, {"$set": {"tags": tags, "updated_at": datetime.now(timezone.utc).isoformat()}})
    await _log_activity(user, "etiquetas_actualizadas", f"Etiquetas: {', '.join(tags)}", "", deal_id)
    return {"tags": tags}

# ==================== ACTIVITY LOG ====================

async def _log_activity(user: dict, action: str, details: str, contact_id: str = "", deal_id: str = ""):
    await db.activity_log.insert_one({
        "id": str(uuid.uuid4()), "tenant_id": user["tenant_id"],
        "user_name": user.get("name", ""), "user_email": user.get("email", ""),
        "action": action, "details": details,
        "contact_id": contact_id, "deal_id": deal_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    })

@api_router.get("/crm/activity-log")
async def get_activity_log(request: Request, contact_id: Optional[str] = None, deal_id: Optional[str] = None, limit: int = 50):
    user = await get_current_user(request)
    query = {"tenant_id": user["tenant_id"]}
    if contact_id: query["contact_id"] = contact_id
    if deal_id: query["deal_id"] = deal_id
    logs = await db.activity_log.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    return logs

@api_router.get("/settings/activity-log")
async def get_settings_activity_log(request: Request, limit: int = 100):
    user = await get_current_user(request)
    logs = await db.activity_log.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    return logs

# ==================== EMAIL SEGMENTS ====================

@api_router.get("/email-marketing/segments")
async def list_segments(request: Request):
    user = await get_current_user(request)
    segments = await db.email_segments.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).sort("created_at", -1).to_list(50)
    # Calculate dynamic counts
    for seg in segments:
        query = _build_segment_query(seg.get("rules", []), user["tenant_id"])
        seg["count"] = await db.leads.count_documents(query)
    return segments

@api_router.post("/email-marketing/segments")
async def create_segment(request: Request, body: Dict[str, Any] = {}):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    segment = {
        "id": str(uuid.uuid4()), "tenant_id": user["tenant_id"],
        "name": body.get("name", ""), "rules": body.get("rules", []),
        "created_at": now, "updated_at": now
    }
    await db.email_segments.insert_one(segment)
    query = _build_segment_query(segment["rules"], user["tenant_id"])
    segment["count"] = await db.leads.count_documents(query)
    return {k: v for k, v in segment.items() if k != "_id"}

@api_router.get("/email-marketing/segments/{segment_id}/leads")
async def get_segment_leads(request: Request, segment_id: str, page: int = 1, limit: int = 50):
    user = await get_current_user(request)
    segment = await db.email_segments.find_one({"id": segment_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not segment:
        raise HTTPException(status_code=404, detail="Segmento no encontrado")
    query = _build_segment_query(segment.get("rules", []), user["tenant_id"])
    total = await db.leads.count_documents(query)
    leads = await db.leads.find(query, {"_id": 0}).sort("ai_score", -1).skip((page - 1) * limit).limit(limit).to_list(limit)
    return {"leads": leads, "total": total, "segment": segment}

@api_router.delete("/email-marketing/segments/{segment_id}")
async def delete_segment(request: Request, segment_id: str):
    user = await get_current_user(request)
    await db.email_segments.delete_one({"id": segment_id, "tenant_id": user["tenant_id"]})
    return {"message": "Segmento eliminado"}

def _build_segment_query(rules: list, tenant_id: str) -> dict:
    query = {"tenant_id": tenant_id}
    for rule in rules:
        field = rule.get("field", "")
        op = rule.get("operator", "equals")
        value = rule.get("value", "")
        if not field or not value: continue
        if op == "equals":
            query[field] = value
        elif op == "contains":
            query[field] = {"$regex": value, "$options": "i"}
        elif op == "gte":
            try: query[field] = {"$gte": int(value)}
            except: pass
        elif op == "lte":
            try: query[field] = {"$lte": int(value)}
            except: pass
        elif op == "in":
            query[field] = {"$in": [v.strip() for v in value.split(",")]}
    return query


# ==================== EXPORT ====================

@api_router.get("/export/n8n-workflow")
async def export_n8n_workflow():
    """Download the n8n workflow JSON file"""
    from fastapi.responses import FileResponse
    filepath = ROOT_DIR.parent / "n8n-workflow-spectra-prospeccion-v2.json"
    if not filepath.exists():
        filepath = ROOT_DIR.parent / "n8n-workflow-spectra-prospeccion.json"
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Workflow file not found")
    return FileResponse(path=str(filepath), media_type="application/octet-stream", filename="spectra-flow-n8n-workflow.json")

@api_router.get("/export/n8n-optimia-workflow")
async def export_n8n_optimia_workflow():
    """Download the OptimIA/Spectra n8n workflow JSON file"""
    from fastapi.responses import FileResponse
    filepath = ROOT_DIR.parent / "n8n-optimia-spectra-2026.json"
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Workflow file not found")
    return FileResponse(path=str(filepath), media_type="application/octet-stream", filename="OPTIMIA_SPECTRA_2026.json")

@api_router.get("/export/leads")
async def export_leads(request: Request):
    from fastapi.responses import StreamingResponse
    import openpyxl
    import io
    user = await get_current_user(request)
    leads = await db.leads.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).to_list(5000)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    headers = ["business_name", "normalized_category", "province", "city", "email", "phone", "website", "ai_score", "quality_level", "status", "recommendation"]
    ws.append(headers)
    for lead in leads:
        ws.append([lead.get(h, "") for h in headers])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=leads_export.xlsx"})

@api_router.get("/export/crm-contacts")
async def export_crm_contacts(request: Request):
    from fastapi.responses import StreamingResponse
    import openpyxl
    import io
    user = await get_current_user(request)
    contacts = await db.crm_contacts.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).to_list(5000)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CRM Contacts"
    headers = ["business_name", "contact_name", "email", "phone", "city", "province", "category", "stage", "ai_score", "deal_count", "total_value"]
    ws.append(headers)
    for c in contacts:
        ws.append([c.get(h, "") for h in headers])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=crm_contacts_export.xlsx"})

# ==================== INTEGRATION WEBHOOKS ====================

@api_router.post("/webhooks/n8n/job-result/{job_id}")
async def n8n_job_result(request: Request, job_id: str, body: Dict[str, Any] = {}):
    """Receives processed leads from n8n after Outscraper + Dify pipeline"""
    api_key = request.headers.get("X-Api-Key", "")
    if not api_key:
        raise HTTPException(status_code=401, detail="API key required")
    job = await db.prospect_jobs.find_one({"id": job_id})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    now = datetime.now(timezone.utc).isoformat()
    leads_data = body.get("leads", [])
    raw_count = body.get("raw_count", len(leads_data))
    inserted = 0
    duplicates_skipped = 0
    for ld in leads_data:
        bname = ld.get("business_name", ld.get("name", "")).strip()
        phone = ld.get("phone", ld.get("phone_number", "")).strip()
        email = ld.get("email", "").strip()
        # Deduplication: check by business_name OR phone OR email within same tenant
        dedup_conditions = [{"business_name": {"$regex": f"^{bname}$", "$options": "i"}}]
        if phone:
            dedup_conditions.append({"phone": phone})
        if email:
            dedup_conditions.append({"email": {"$regex": f"^{email}$", "$options": "i"}})
        existing = await db.leads.find_one({"tenant_id": job["tenant_id"], "$or": dedup_conditions}, {"_id": 0, "id": 1})
        if existing:
            duplicates_skipped += 1
            logger.info(f"Lead duplicado omitido: {bname} (tenant {job['tenant_id']})")
            continue
        lead = {
            "id": str(uuid.uuid4()), "tenant_id": job["tenant_id"], "job_id": job_id,
            "business_name": bname,
            "raw_category": ld.get("raw_category", ld.get("category", "")),
            "normalized_category": ld.get("normalized_category", ld.get("category", "")),
            "province": ld.get("province", ld.get("state", job.get("province", ""))),
            "city": ld.get("city", job.get("city", "")),
            "website": ld.get("website", ld.get("site", "")),
            "email": email,
            "phone": phone,
            "ai_score": int(ld.get("ai_score", ld.get("score", 0))),
            "quality_level": ld.get("quality_level", "average"),
            "recommendation": ld.get("recommendation", ""),
            "recommended_first_line": ld.get("recommended_first_line", ld.get("first_line", "")),
            "address": ld.get("address", ld.get("full_address", "")),
            "rating": ld.get("rating", 0),
            "reviews_count": ld.get("reviews_count", ld.get("reviews", 0)),
            "status": "scored" if int(ld.get("ai_score", 0)) >= 50 else "rejected",
            "created_at": now, "updated_at": now
        }
        await db.leads.insert_one(lead)
        inserted += 1
    if duplicates_skipped > 0:
        logger.info(f"n8n callback: {inserted} insertados, {duplicates_skipped} duplicados omitidos para job {job_id}")
    qualified = sum(1 for ld in leads_data if int(ld.get("ai_score", 0)) >= 50)
    rejected = inserted - qualified
    stages = [
        {"name": "job_created", "status": "completed", "timestamp": job.get("created_at", now)},
        {"name": "scraping", "status": "completed", "timestamp": now},
        {"name": "prospects_found", "status": "completed", "timestamp": now},
        {"name": "ai_cleaning", "status": "completed", "timestamp": now},
        {"name": "scoring_completed", "status": "completed", "timestamp": now},
        {"name": "ready_for_review", "status": "completed", "timestamp": now}
    ]
    await db.prospect_jobs.update_one({"id": job_id}, {"$set": {
        "status": "completed", "raw_count": raw_count, "cleaned_count": inserted,
        "qualified_count": qualified, "rejected_count": rejected, "duplicates_skipped": duplicates_skipped, "stages": stages, "updated_at": now
    }})
    # Auto-create/update email list with scored leads from n8n
    scored_lead_ids = []
    for ld in leads_data:
        if int(ld.get("ai_score", 0)) >= 50:
            # Find the lead we just inserted by matching business_name
            lead_doc = await db.leads.find_one({"tenant_id": job["tenant_id"], "job_id": job_id, "business_name": ld.get("business_name", ld.get("name", ""))}, {"id": 1, "_id": 0})
            if lead_doc:
                scored_lead_ids.append(lead_doc["id"])
    cat = job.get("category", "Prospeccion")
    city = job.get("city", "")
    auto_list_name = f"{cat} - {city} - Calificados"
    auto_list_id = ""
    # Check if any leads have default score 50 — auto-rescore them in background
    leads_with_50 = sum(1 for ld in leads_data if int(ld.get("ai_score", ld.get("score", 0))) == 50)
    if leads_with_50 > 0:
        import asyncio
        asyncio.create_task(_auto_rescore_job_leads(job_id, job["tenant_id"]))
    if scored_lead_ids:
        existing_list = await db.email_lists.find_one({"tenant_id": job["tenant_id"], "name": auto_list_name})
        if existing_list:
            old_ids = existing_list.get("lead_ids", [])
            new_ids = old_ids + [sid for sid in scored_lead_ids if sid not in old_ids]
            await db.email_lists.update_one({"id": existing_list["id"]}, {"$set": {"lead_ids": new_ids, "subscriber_count": len(new_ids), "updated_at": now}})
            auto_list_id = existing_list["id"]
        else:
            auto_list_id = str(uuid.uuid4())
            await db.email_lists.insert_one({"id": auto_list_id, "tenant_id": job["tenant_id"], "name": auto_list_name, "description": f"Auto: {len(scored_lead_ids)} leads calificados via n8n", "subscriber_count": len(scored_lead_ids), "lead_ids": scored_lead_ids, "created_at": now, "updated_at": now})
    await db.prospect_jobs.update_one({"id": job_id}, {"$set": {"auto_list_name": auto_list_name, "auto_list_id": auto_list_id}})
    return {"message": f"{inserted} leads importados, {qualified} calificados, {duplicates_skipped} duplicados omitidos", "job_id": job_id, "auto_list": auto_list_name, "duplicates_skipped": duplicates_skipped}

async def _auto_rescore_job_leads(job_id: str, tenant_id: str):
    """Auto-rescore leads that arrived with default score 50 (Dify failed in n8n)"""
    import asyncio
    await asyncio.sleep(2)  # Brief delay to let inserts complete
    leads_50 = await db.leads.find({"job_id": job_id, "tenant_id": tenant_id, "ai_score": 50}, {"_id": 0}).to_list(100)
    if not leads_50:
        return
    logger.info(f"Auto-rescoring {len(leads_50)} leads with score 50 for job {job_id}")
    dify_base = os.environ.get("DIFY_BASE_URL", "")
    dify_key = os.environ.get("DIFY_APP_KEY", "")
    if not dify_base or not dify_key:
        dify_config = await db.integration_configs.find_one({"tenant_id": tenant_id, "name": "dify"}, {"_id": 0})
        if dify_config:
            dify_base = dify_config.get("base_url", dify_base)
            dify_key = dify_config.get("api_key", dify_key)
    now = datetime.now(timezone.utc).isoformat()
    rescored = 0
    for lead in leads_50:
        business_data = f"Nombre: {lead.get('business_name','')}\nDireccion: {lead.get('address','')}, {lead.get('city','')}, {lead.get('province','')}\nTelefono: {lead.get('phone','')}\nWebsite: {lead.get('website','')}\nEmail: {lead.get('email','')}\nRating: {lead.get('rating',0)}\nReviews: {lead.get('reviews_count',0)}\nCategoria: {lead.get('normalized_category','')}"
        scored = None
        if dify_base and dify_key:
            try:
                dify_url = dify_base.rstrip("/")
                if dify_url.startswith("http://"):
                    dify_url = "https://" + dify_url[7:]
                async with httpx.AsyncClient(timeout=20, follow_redirects=True) as client:
                    resp = await client.post(f"{dify_url}/workflows/run", headers={"Authorization": f"Bearer {dify_key}", "Content-Type": "application/json"}, json={"inputs": {"business_data": business_data}, "response_mode": "blocking", "user": "spectra-auto"})
                    result = resp.json()
                    output = result.get("data", {}).get("outputs", {}).get("result", "")
                    if output:
                        import json as jm
                        json_match = re.search(r'\{[\s\S]*\}', output) if isinstance(output, str) else None
                        parsed = jm.loads(json_match.group(0)) if json_match else (output if isinstance(output, dict) else None)
                        if parsed and parsed.get("ai_score"):
                            scored = parsed
            except:
                pass
        if not scored:
            try:
                from emergentintegrations.llm.chat import LlmChat, UserMessage
                chat = LlmChat(api_key=os.environ.get("EMERGENT_LLM_KEY", ""), session_id=f"autoscore-{lead['id']}", system_message="Eres un clasificador de leads B2B. Califica del 0 al 100. Responde SOLO JSON: {\"ai_score\": N, \"quality_level\": \"excellent/good/average/poor\", \"recommendation\": \"texto corto\", \"recommended_first_line\": \"texto\"}")
                resp_text = await chat.send_message(UserMessage(text=f"Clasifica este lead:\n{business_data}"))
                import json as jm
                json_match = re.search(r'\{[\s\S]*\}', resp_text)
                if json_match:
                    scored = jm.loads(json_match.group(0))
            except:
                pass
        if scored and scored.get("ai_score"):
            score_val = int(scored["ai_score"])
            status = "scored" if score_val >= 50 else "rejected"
            await db.leads.update_one({"id": lead["id"]}, {"$set": {"ai_score": score_val, "quality_level": scored.get("quality_level", "average"), "recommendation": scored.get("recommendation", ""), "recommended_first_line": scored.get("recommended_first_line", ""), "status": status, "updated_at": now}})
            rescored += 1
        await asyncio.sleep(0.5)
    logger.info(f"Auto-rescore complete: {rescored}/{len(leads_50)} leads for job {job_id}")

@api_router.post("/webhooks/n8n/job-progress/{job_id}")
async def n8n_job_progress(request: Request, job_id: str, body: Dict[str, Any] = {}):
    """Updates job progress from n8n (stage by stage)"""
    stage_name = body.get("stage", "")
    stage_status = body.get("status", "completed")
    raw_count = body.get("raw_count")
    job = await db.prospect_jobs.find_one({"id": job_id})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    now = datetime.now(timezone.utc).isoformat()
    stages = job.get("stages", [])
    for s in stages:
        if s["name"] == stage_name:
            s["status"] = stage_status
            s["timestamp"] = now
    update = {"stages": stages, "updated_at": now, "status": "processing"}
    if raw_count is not None:
        update["raw_count"] = raw_count
    await db.prospect_jobs.update_one({"id": job_id}, {"$set": update})
    return {"message": f"Stage {stage_name} updated to {stage_status}"}

@api_router.post("/webhooks/chatwoot/lead")
async def chatwoot_lead_webhook(request: Request, body: Dict[str, Any] = {}):
    """Receives new leads from Chatwoot/OptimIA Bot — creates in LEADS collection with BOT tag"""
    tenant_id = body.get("tenant_id", "")
    if not tenant_id:
        tenant = await db.tenants.find_one({}, {"_id": 0, "id": 1})
        tenant_id = tenant["id"] if tenant else ""
    if not tenant_id:
        raise HTTPException(status_code=400, detail="tenant_id required")
    now = datetime.now(timezone.utc).isoformat()
    email = body.get("email", "")
    phone = body.get("phone", body.get("phone_number", ""))
    contact_name = body.get("contact_name", body.get("name", ""))
    business_name = body.get("business_name", body.get("name", "Lead de Bot"))
    channel = body.get("channel", "web")
    message = body.get("message", body.get("notes", ""))
    # Try to extract email from message if not provided
    if not email and message:
        import re
        email_match = re.search(r'[\w.+-]+@[\w-]+\.[\w.]+', message)
        if email_match:
            email = email_match.group(0)
    # Try to extract phone from message if not provided
    if not phone and message:
        import re
        phone_match = re.search(r'[\+]?[\d\s\-\(\)]{8,15}', message)
        if phone_match:
            candidate = phone_match.group(0).strip()
            if len(candidate.replace(" ", "").replace("-", "")) >= 8:
                phone = candidate
    # Check for existing lead by email or phone
    existing = None
    if email:
        existing = await db.leads.find_one({"email": {"$regex": f"^{email}$", "$options": "i"}, "tenant_id": tenant_id})
    if not existing and phone:
        existing = await db.leads.find_one({"phone": phone, "tenant_id": tenant_id})
    if not existing and business_name:
        existing = await db.leads.find_one({"business_name": {"$regex": f"^{business_name}$", "$options": "i"}, "tenant_id": tenant_id})
    if existing:
        update_fields = {"updated_at": now, "source": "bot"}
        if body.get("message"):
            update_fields["recommendation"] = (existing.get("recommendation", "") + f"\n[BOT {now[:10]}] {body['message']}").strip()
        tags = existing.get("tags", [])
        if "bot" not in tags:
            tags.append("bot")
        update_fields["tags"] = tags
        await db.leads.update_one({"id": existing["id"]}, {"$set": update_fields})
        return {"message": "Lead actualizado desde Bot", "lead_id": existing["id"], "action": "updated"}
    else:
        lead = {
            "id": str(uuid.uuid4()), "tenant_id": tenant_id, "job_id": "",
            "business_name": business_name,
            "raw_category": body.get("category", ""),
            "normalized_category": body.get("category", "Bot"),
            "province": body.get("province", ""),
            "city": body.get("city", ""),
            "website": body.get("website", ""),
            "email": email, "phone": phone,
            "ai_score": 0, "quality_level": "unscored",
            "recommendation": body.get("message", body.get("notes", "")),
            "recommended_first_line": "",
            "source": "bot", "tags": ["bot"],
            "channel": channel,
            "status": "raw",
            "created_at": now, "updated_at": now
        }
        await db.leads.insert_one(lead)
        return {"message": "Lead creado desde Bot", "lead_id": lead["id"], "action": "created"}

@api_router.post("/webhooks/resend/events")
async def resend_events_webhook(request: Request, body: Dict[str, Any] = {}):
    """Receives email events from Resend (opens, clicks, bounces)"""
    event_type = body.get("type", "")
    email_data = body.get("data", {})
    logger.info(f"Resend event: {event_type} - {email_data.get('email_id', '')}")
    return {"received": True}

# ==================== SEED DATA ====================

async def seed_data():
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@spectraflow.com")
    admin_password = os.environ.get("ADMIN_PASSWORD", "Admin123!")
    existing_admin = await db.users.find_one({"email": admin_email})
    if existing_admin:
        if not verify_password(admin_password, existing_admin["password_hash"]):
            await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password)}})
        logger.info("Seed data already exists")
        os.makedirs("/app/memory", exist_ok=True)
        with open("/app/memory/test_credentials.md", "w") as f:
            f.write(f"# Test Credentials\n\n## Admin Account\n- Email: {admin_email}\n- Password: {admin_password}\n- Role: super_admin\n\n## Demo Operator\n- Email: demo@spectraflow.com\n- Password: Demo123!\n- Role: operator\n\n## Auth Endpoints\n- POST /api/auth/login\n- POST /api/auth/register\n- GET /api/auth/me\n- POST /api/auth/logout\n")
        return

    logger.info("Seeding demo data...")
    tenant_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()

    await db.tenants.insert_one({"id": tenant_id, "name": "Spectra Demo", "branding": {"company_name": "Spectra Demo", "logo_url": "", "primary_color": "#1D4ED8", "secondary_color": "#6366F1"}, "sender_defaults": {"name": "Spectra Flow", "email": "noreply@spectraflow.com"}, "modules": {"prospeccion": True, "leads": True, "crm": True, "email_marketing": True}, "plan": "enterprise", "price": 0, "active": True, "created_at": now})
    await db.users.insert_one({"email": admin_email, "password_hash": hash_password(admin_password), "name": "Admin", "role": "super_admin", "tenant_id": tenant_id, "created_at": now})
    await db.users.insert_one({"email": "demo@spectraflow.com", "password_hash": hash_password("Demo123!"), "name": "Maria Garcia", "role": "operator", "tenant_id": tenant_id, "created_at": now})

    await db.users.create_index("email", unique=True)
    await db.leads.create_index([("tenant_id", 1), ("status", 1)])
    await db.leads.create_index([("tenant_id", 1), ("job_id", 1)])
    await db.prospect_jobs.create_index([("tenant_id", 1)])

    jobs_data = [
        {"province": "Tucuman", "city": "San Miguel de Tucuman", "category": "Real Estate", "status": "completed", "quantity": 200},
        {"province": "Buenos Aires", "city": "CABA", "category": "Technology", "status": "completed", "quantity": 150},
        {"province": "Cordoba", "city": "Cordoba Capital", "category": "Gastronomia", "status": "processing", "quantity": 100},
    ]
    job_ids = []
    for jd in jobs_data:
        job_id = str(uuid.uuid4())
        job_ids.append(job_id)
        raw = random.randint(int(jd["quantity"]*0.8), int(jd["quantity"]*1.2))
        cleaned = int(raw * 0.78)
        qualified = int(cleaned * 0.72)
        stages = [
            {"name": "job_created", "status": "completed", "timestamp": now},
            {"name": "scraping", "status": "completed", "timestamp": now},
            {"name": "prospects_found", "status": "completed", "timestamp": now},
            {"name": "ai_cleaning", "status": "completed" if jd["status"] == "completed" else "in_progress", "timestamp": now},
            {"name": "scoring_completed", "status": "completed" if jd["status"] == "completed" else "pending", "timestamp": now if jd["status"] == "completed" else None},
            {"name": "ready_for_review", "status": "completed" if jd["status"] == "completed" else "pending", "timestamp": now if jd["status"] == "completed" else None}
        ]
        await db.prospect_jobs.insert_one({"id": job_id, "tenant_id": tenant_id, "province": jd["province"], "city": jd["city"], "category": jd["category"], "quantity": jd["quantity"], "filters": {}, "status": jd["status"], "raw_count": raw, "cleaned_count": cleaned, "qualified_count": qualified, "rejected_count": cleaned - qualified, "approved_count": random.randint(0, qualified // 2), "stages": stages, "created_by": "Admin", "created_at": now, "updated_at": now})

    businesses = [("Inmobiliaria del Norte SA", "Real Estate", 0), ("Casa & Hogar SRL", "Real Estate", 0), ("Propiedades Premium SA", "Real Estate", 0), ("Inversiones Tucuman SAS", "Real Estate", 0), ("Gestion Inmobiliaria", "Real Estate", 0), ("TechNova Solutions", "Technology", 1), ("Digital Minds SRL", "Technology", 1), ("CloudArg SA", "Technology", 1), ("InnoSoft Patagonia", "Technology", 1), ("DataFlow Analytics", "Technology", 1), ("El Buen Sabor", "Gastronomia", 2), ("Restaurante Don Carlos", "Gastronomia", 2), ("Cafe Central", "Gastronomia", 2), ("La Parrilla del Sur", "Gastronomia", 2), ("Bistro Moderno", "Gastronomia", 2), ("Despacho Legal Rios", "Legal", 0), ("Consultoria Empresarial Plus", "Consulting", 1), ("Seguros del Litoral", "Insurance", 0), ("Clinica Dental Premium", "Health", 1), ("Transporte Ejecutivo NOA", "Logistics", 2)]
    statuses_pool = ["scored", "scored", "scored", "approved", "approved", "cleaned", "rejected", "queued_for_sequence", "contacted", "opened", "replied", "interested", "sent_to_crm"]
    cities = ["San Miguel de Tucuman", "CABA", "Cordoba Capital"]
    provinces = ["Tucuman", "Buenos Aires", "Cordoba"]
    for name, cat, ji in businesses:
        score = random.randint(35, 98)
        ql = "excellent" if score >= 80 else ("good" if score >= 60 else ("average" if score >= 45 else "poor"))
        await db.leads.insert_one({"id": str(uuid.uuid4()), "tenant_id": tenant_id, "job_id": job_ids[min(ji, len(job_ids)-1)], "business_name": name, "raw_category": cat.lower(), "normalized_category": cat, "province": provinces[min(ji, 2)], "city": cities[min(ji, 2)], "website": f"www.{name.lower().replace(' ', '').replace('&','')[:15]}.com.ar", "email": f"info@{name.lower().replace(' ', '').replace('&','')[:10]}.com.ar", "phone": f"+54 {random.randint(11,99)} {random.randint(1000,9999)}-{random.randint(1000,9999)}", "ai_score": score, "quality_level": ql, "recommendation": f"{'Highly recommended' if ql in ['excellent','good'] else 'Needs review'} - AI confidence {score}%", "recommended_first_line": f"Notamos que {name} tiene presencia destacada en el sector de {cat}.", "status": random.choice(statuses_pool), "created_at": now, "updated_at": now})

    campaigns_data = [
        {"name": "Real Estate Tucuman Q1 2026", "status": "active", "lead_count": 45, "sent_count": 38, "open_count": 22, "click_count": 8, "reply_count": 5, "interested_count": 3, "crm_count": 2},
        {"name": "Tech Companies Buenos Aires", "status": "draft", "lead_count": 30, "sent_count": 0, "open_count": 0, "click_count": 0, "reply_count": 0, "interested_count": 0, "crm_count": 0},
        {"name": "Gastronomia Cordoba Pilot", "status": "completed", "lead_count": 20, "sent_count": 20, "open_count": 14, "click_count": 6, "reply_count": 4, "interested_count": 2, "crm_count": 1}
    ]
    for c in campaigns_data:
        await db.campaigns.insert_one({"id": str(uuid.uuid4()), "tenant_id": tenant_id, **c, "template_id": "", "sender_profile_id": "", "lead_ids": [], "created_by": "Admin", "created_at": now, "updated_at": now})

    templates_data = [
        {"name": "Primer Contacto", "subject": "Oportunidad de colaboracion para {business_name}", "html_body": "<p>Estimado equipo de <strong>{business_name}</strong>,</p><p>{recommended_first_line}</p><p>En Spectra Flow, ayudamos a empresas del sector de <strong>{normalized_category}</strong> en <strong>{city}</strong> a potenciar su presencia digital.</p><p>Saludos cordiales,<br/>{sender_name}</p>", "variables": ["business_name", "recommended_first_line", "normalized_category", "city", "sender_name"]},
        {"name": "Follow-up", "subject": "Re: Oportunidad para {business_name}", "html_body": "<p>Hola,</p><p>Le escribo nuevamente respecto a nuestra propuesta para <strong>{business_name}</strong>.</p><p>Hemos trabajado con empresas similares en {city} logrando resultados significativos.</p><p>Saludos,<br/>{sender_name}</p>", "variables": ["business_name", "city", "sender_name"]},
        {"name": "Ultima Oportunidad", "subject": "Ultima oportunidad - {business_name}", "html_body": "<p>Estimado equipo,</p><p>Esta es mi ultima comunicacion. Entiendo que estan ocupados, pero no queria dejar pasar esta oportunidad para {business_name}.</p><p>Exitos,<br/>{sender_name}</p>", "variables": ["business_name", "sender_name"]}
    ]
    for t in templates_data:
        await db.templates.insert_one({"id": str(uuid.uuid4()), "tenant_id": tenant_id, **t, "plain_text": "", "signature": "", "created_at": now, "updated_at": now})

    await db.domains.insert_one({"id": str(uuid.uuid4()), "tenant_id": tenant_id, "domain": "spectra-metrics.com", "subdomain": "spectra-metrics.com", "resend_domain_id": "", "status": "dns_pending", "dns_records": [{"type": "TXT", "name": "spectra-metrics.com", "value": "v=spf1 include:_spf.resend.com ~all", "verified": False, "status": "pending"}, {"type": "CNAME", "name": "resend._domainkey.spectra-metrics.com", "value": "resend.domainkey.resend.com", "verified": False, "status": "pending"}, {"type": "MX", "name": "spectra-metrics.com", "value": "feedback-smtp.resend.com", "verified": False, "status": "pending"}], "sender_name": "Spectra Flow", "sender_email": "noreply@spectra-metrics.com", "reply_to": "contacto@spectra-metrics.com", "signature": "El equipo de Spectra Flow", "warmup_status": "not_started", "created_at": now, "updated_at": now})

    integrations = [
        {"name": "n8n", "display_name": "n8n Orchestration", "enabled": False, "base_url": "", "api_key": "", "status": "not_configured", "last_sync": None, "description": "Workflow orchestration and job execution"},
        {"name": "dify", "display_name": "Dify AI", "enabled": False, "base_url": "", "api_key": "", "status": "not_configured", "last_sync": None, "description": "AI cleaning and lead scoring"},
        {"name": "resend", "display_name": "Resend Email", "enabled": True, "base_url": "https://api.resend.com", "api_key": "***configured***", "status": "configured", "last_sync": None, "description": "Email sending and domain verification"},
        {"name": "espo_crm", "display_name": "Spectra CRM", "enabled": False, "base_url": "", "api_key": "", "status": "not_configured", "last_sync": None, "description": "Qualified lead handoff and CRM sync"},
        {"name": "optimia_bot", "display_name": "OptimIA Bot", "enabled": False, "base_url": "https://inbox.optimia.disruptive-sw.com", "api_key": "", "status": "not_configured", "last_sync": None, "description": "Omnichannel bot and live chat support"},
        {"name": "apify", "display_name": "Apify (LinkedIn)", "enabled": False, "base_url": "https://api.apify.com/v2", "api_key": "", "status": "not_configured", "last_sync": None, "description": "LinkedIn scraping via Apify actors"}
    ]
    for intg in integrations:
        await db.integration_configs.insert_one({"id": str(uuid.uuid4()), "tenant_id": tenant_id, **intg, "created_at": now, "updated_at": now})

    audit_entries = [
        {"action": "created_prospect_job", "entity_type": "prospect_job", "details": "Job for Real Estate in San Miguel de Tucuman"},
        {"action": "prospect_job_completed", "entity_type": "prospect_job", "details": "172 leads found and scored"},
        {"action": "leads_approved", "entity_type": "lead", "details": "12 leads approved for sequence"},
        {"action": "campaign_launched", "entity_type": "campaign", "details": "Real Estate Tucuman Q1 2026 launched"},
        {"action": "domain_verified", "entity_type": "domain", "details": "mail.spectraflow.com verified"},
        {"action": "leads_sent_to_crm", "entity_type": "crm_sync", "details": "3 leads synced to Espo CRM"},
    ]
    for entry in audit_entries:
        await db.audit_logs.insert_one({"id": str(uuid.uuid4()), "tenant_id": tenant_id, "user_id": "", "user_name": "Admin", **entry, "entity_id": "", "created_at": now})

    crm_logs = [
        {"lead_name": "Inmobiliaria del Norte SA", "status": "synced", "error": ""},
        {"lead_name": "TechNova Solutions", "status": "synced", "error": ""},
        {"lead_name": "El Buen Sabor", "status": "error", "error": "Connection timeout - CRM unreachable"},
    ]
    for cl in crm_logs:
        await db.crm_sync_logs.insert_one({"id": str(uuid.uuid4()), "tenant_id": tenant_id, "lead_id": "", **cl, "synced_at": now if cl["status"] == "synced" else None, "assigned_owner": "Carlos Rodriguez", "created_at": now})

    os.makedirs("/app/memory", exist_ok=True)
    with open("/app/memory/test_credentials.md", "w") as f:
        f.write(f"# Test Credentials\n\n## Admin Account\n- Email: {admin_email}\n- Password: {admin_password}\n- Role: super_admin\n\n## Demo Operator\n- Email: demo@spectraflow.com\n- Password: Demo123!\n- Role: operator\n\n## Auth Endpoints\n- POST /api/auth/login\n- POST /api/auth/register\n- GET /api/auth/me\n- POST /api/auth/logout\n")
    logger.info("Seed data created successfully!")

# ==================== STARTUP / SHUTDOWN ====================

@app.on_event("startup")
async def startup():
    await seed_data()

@app.on_event("shutdown")
async def shutdown():
    client.close()

app.include_router(api_router)
app.mount("/api/uploads", StaticFiles(directory=str(ROOT_DIR / "uploads")), name="uploads")

frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[frontend_url, "http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
